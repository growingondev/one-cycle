from __future__ import annotations

import argparse
import json
import subprocess
import sys
import time
import urllib.error
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from evaluation.dataset_resolver import (
    DEFAULT_SHEET_NAME,
    default_manifest_path,
    default_result_path,
    resolve_dataset_xlsx,
    resolve_evaluation_documents,
)


# ============================================================
# 기본 설정
# ============================================================

EVALUATION_DB_NAME = "one_cycle_evaluation_tmp"

DEFAULT_API_BASE_URL = "http://127.0.0.1:18000"
DEFAULT_ENDPOINT = "/api/chat"

DEFAULT_TIMEOUT_SECONDS = 600
DEFAULT_RETRY_COUNT = 1

DEFAULT_RUN_NUMBER = "001"


# ============================================================
# Git 정보
# ============================================================


def get_git_info() -> tuple[str, str]:
    """
    현재 평가 실행 시점의 Git branch / commit을 기록한다.
    """

    try:
        branch = subprocess.check_output(
            [
                "git",
                "branch",
                "--show-current",
            ],
            text=True,
            stderr=subprocess.DEVNULL,
        ).strip()

    except Exception:
        branch = ""

    try:
        commit = subprocess.check_output(
            [
                "git",
                "rev-parse",
                "--short",
                "HEAD",
            ],
            text=True,
            stderr=subprocess.DEVNULL,
        ).strip()

    except Exception:
        commit = ""

    return (
        branch,
        commit,
    )


# ============================================================
# Excel
# ============================================================


def find_columns(
    ws,
) -> dict[str, int]:
    """
    Excel 1행의 column 이름을 읽어
    {column_name: column_index} 형태로 반환한다.
    """

    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def set_optional_cell(
    *,
    ws,
    columns: dict[str, int],
    row: int,
    name: str,
    value: Any,
) -> None:
    """
    평가셋에 해당 column이 존재할 때만 값을 저장한다.

    데이터셋마다 부가 column 구성이 조금 달라도
    evaluate_rag.py가 깨지지 않도록 한다.
    """

    column = columns.get(
        name
    )

    if column is None:
        return

    ws.cell(
        row=row,
        column=column,
        value=value,
    )


# ============================================================
# question_id 선택
# ============================================================


def parse_question_ids(
    value: str | None,
) -> set[str] | None:
    """
    예:
        --question-ids Q001,Q003,Q010

    ->
        {"Q001", "Q003", "Q010"}
    """

    if value is None:
        return None

    question_ids = {
        item.strip().upper()
        for item in value.split(",")
        if item.strip()
    }

    return (
        question_ids
        or None
    )


# ============================================================
# API
# ============================================================


def build_api_url(
    *,
    api_base_url: str,
    endpoint: str,
) -> str:

    return (
        api_base_url.rstrip("/")
        + "/"
        + endpoint.lstrip("/")
    )


def http_post_json(
    *,
    url: str,
    payload: dict[str, Any],
    timeout: int,
) -> dict[str, Any]:
    """
    POST /api/chat 호출.

    외부 requests 패키지를 추가하지 않고
    Python 표준 urllib만 사용한다.
    """

    body = json.dumps(
        payload,
        ensure_ascii=False,
    ).encode(
        "utf-8"
    )

    request = urllib.request.Request(
        url=url,
        data=body,
        method="POST",
        headers={
            "Content-Type":
                "application/json",
            "Accept":
                "application/json",
        },
    )

    try:
        with urllib.request.urlopen(
            request,
            timeout=timeout,
        ) as response:

            raw = (
                response
                .read()
                .decode(
                    "utf-8"
                )
            )

            result = json.loads(
                raw
            )

    except urllib.error.HTTPError as exc:

        error_body = (
            exc.read()
            .decode(
                "utf-8",
                errors="replace",
            )
        )

        raise RuntimeError(
            f"HTTP {exc.code} 오류\n"
            f"URL: {url}\n"
            f"응답: {error_body}"
        ) from exc

    except urllib.error.URLError as exc:

        raise RuntimeError(
            "API 서버 연결 실패: "
            f"{exc.reason}\n"
            f"URL: {url}"
        ) from exc

    except TimeoutError as exc:

        raise RuntimeError(
            f"API 요청 Timeout: {timeout}s\n"
            f"URL: {url}"
        ) from exc

    except json.JSONDecodeError as exc:

        raise RuntimeError(
            "API 응답이 JSON 형식이 아닙니다.\n"
            f"URL: {url}"
        ) from exc

    if not isinstance(
        result,
        dict,
    ):
        raise RuntimeError(
            "API 응답 형식이 dict가 아닙니다."
        )

    return result


def request_chat_with_retry(
    *,
    api_url: str,
    announcement_id: int,
    question: str,
    timeout: int,
    retry_count: int,
) -> dict[str, Any]:
    """
    /api/chat 요청 실패 시 retry_count 만큼 재시도한다.

    retry=1
        최초 1회 + 재시도 1회
        총 2회
    """

    payload = {
        "announcementId":
            announcement_id,
        "question":
            question,
    }

    last_error: Exception | None = (
        None
    )

    total_attempts = (
        retry_count
        + 1
    )

    for attempt in range(
        1,
        total_attempts + 1,
    ):

        try:
            return http_post_json(
                url=api_url,
                payload=payload,
                timeout=timeout,
            )

        except Exception as exc:

            last_error = exc

            print(
                "  요청 실패 "
                f"({attempt}/{total_attempts}): "
                f"{exc}"
            )

            if (
                attempt
                < total_attempts
            ):
                time.sleep(
                    2
                )

    assert (
        last_error
        is not None
    )

    raise last_error


# ============================================================
# Evidence 저장
# ============================================================


def normalize_evidence(
    value: Any,
) -> list[dict[str, Any]]:
    """
    /api/chat evidence가 list[dict]인지 확인한다.
    """

    if not isinstance(
        value,
        list,
    ):
        return []

    return [
        item
        for item in value
        if isinstance(
            item,
            dict,
        )
    ]


def serialize_chunk_ids(
    evidence: list[
        dict[str, Any]
    ],
) -> str:

    chunk_ids: list[str] = []

    for item in evidence:

        chunk_id = item.get(
            "chunkId",
            item.get(
                "chunk_id"
            ),
        )

        if chunk_id is not None:
            chunk_ids.append(
                str(
                    chunk_id
                )
            )

    return ", ".join(
        chunk_ids
    )


def serialize_contexts(
    evidence: list[
        dict[str, Any]
    ],
) -> str:
    """
    evaluate_metrics.py의
    split_retrieved_contexts()가 다시 분리할 수 있도록

    [rank=1 | ...]
    context

    ---

    [rank=2 | ...]
    context

    형식으로 저장한다.
    """

    contexts: list[str] = []

    for rank, item in enumerate(
        evidence,
        start=1,
    ):

        chunk_id = item.get(
            "chunkId",
            item.get(
                "chunk_id"
            ),
        )

        section_title = (
            item.get(
                "sectionTitle",
                item.get(
                    "section_title"
                ),
            )
        )

        content = str(
            item.get(
                "content"
            )
            or ""
        )

        score = item.get(
            "score"
        )

        parts = [
            f"rank={rank}"
        ]

        if chunk_id is not None:

            parts.append(
                f"chunkId={chunk_id}"
            )

        if section_title:

            parts.append(
                "section="
                f"{section_title}"
            )

        if score is not None:

            parts.append(
                f"score={score}"
            )

        header = (
            "["
            + " | ".join(
                parts
            )
            + "]"
        )

        contexts.append(
            (
                f"{header}\n"
                f"{content}"
            ).strip()
        )

    return (
        "\n\n---\n\n".join(
            contexts
        )
    )


def serialize_evidence_json(
    evidence: list[
        dict[str, Any]
    ],
) -> str:

    return json.dumps(
        evidence,
        ensure_ascii=False,
    )


# ============================================================
# 평가 DB 확인
# ============================================================


def assert_evaluation_database() -> None:
    """
    평가용 문서 등록 / 문서처리가
    운영 DB에서 실행되는 사고를 방지한다.

    여기서는 evaluate_rag.py가 직접 사용하는
    Backend DB 연결을 확인한다.

    실제 /api/chat Backend와 RAG 컨테이너 역시
    one_cycle_evaluation_tmp를 바라보도록
    별도로 환경 설정되어 있어야 한다.
    """

    from sqlalchemy import text

    from backend.app.core.config import (
        settings,
    )

    from backend.app.db.session import (
        SessionLocal,
    )

    configured_database = (
        settings.postgres_db
    )

    if (
        configured_database
        != EVALUATION_DB_NAME
    ):
        raise RuntimeError(
            "평가 작업은 평가 DB에서만 "
            "실행할 수 있습니다. "
            f"configured="
            f"{configured_database}, "
            f"required="
            f"{EVALUATION_DB_NAME}"
        )

    with SessionLocal() as db:

        actual_database = (
            db.execute(
                text(
                    "SELECT "
                    "current_database()"
                )
            )
            .scalar_one()
        )

    if (
        actual_database
        != EVALUATION_DB_NAME
    ):
        raise RuntimeError(
            "실제 연결 DB가 "
            "평가 DB가 아닙니다. "
            f"actual="
            f"{actual_database}, "
            f"required="
            f"{EVALUATION_DB_NAME}"
        )

def activate_evaluation_collection(
    *,
    manifest: dict[str, Any],
) -> dict[str, Any]:
    """
    현재 평가하려는 dataset의 CollectionRun을
    Evaluation DB의 active collection으로 전환한다.

    evaluate_rag.py는 manifest에 저장된
    collection_run_id를 기준으로 자동 전환한다.

    예:
        BD manifest -> collection_run_id=6
        GC manifest -> collection_run_id=7

    따라서 BD/GC/DH 등 평가셋을 변경할 때
    system_state.active_collection_run_id를
    수동으로 변경할 필요가 없다.
    """

    assert_evaluation_database()

    collection_run_id_raw = (
        manifest.get(
            "collection_run_id"
        )
    )

    if collection_run_id_raw is None:
        raise RuntimeError(
            "manifest에 collection_run_id가 없습니다."
        )

    try:
        collection_run_id = int(
            collection_run_id_raw
        )

    except (
        TypeError,
        ValueError,
    ) as exc:
        raise RuntimeError(
            "manifest의 collection_run_id가 "
            "올바른 정수가 아닙니다: "
            f"{collection_run_id_raw}"
        ) from exc

    if collection_run_id <= 0:
        raise RuntimeError(
            "manifest의 collection_run_id는 "
            "1 이상이어야 합니다: "
            f"{collection_run_id}"
        )

    from backend.app.services.collection_publish_service import (
        publish_collection_run,
    )

    result = publish_collection_run(
        collection_run_id
    )

    active_collection_run_id = int(
        result[
            "active_collection_run_id"
        ]
    )

    if (
        active_collection_run_id
        != collection_run_id
    ):
        raise RuntimeError(
            "평가 Collection 활성화에 실패했습니다. "
            f"expected={collection_run_id}, "
            f"actual={active_collection_run_id}"
        )

    return result

# ============================================================
# 실제 문서 Pipeline 준비
# ============================================================


def prepare_actual_pipeline(
    *,
    dataset_id: str,
    xlsx_path: Path,
    sheet_name: str,
    manifest_path: Path,
) -> dict[str, Any]:
    """
    평가 원본 문서를 실제 서비스와 동일한
    문서처리 Pipeline에 태운다.

    평가 원본 HWP/HWPX
        ↓
    register_evaluation_dataset()
        ↓
    Evaluation DB
        ↓
    process_and_publish_evaluation_collection()
        ↓
    Document Worker
        ↓
    Parser
        ↓
    Normalizer
        ↓
    Structure / Verification
        ↓
    Chunking
        ↓
    Embedding
        ↓
    Backend Persistence
        ↓
    Publish
    """

    from backend.app.services.evaluation_service import (
        EvaluationDocumentInput,
        register_evaluation_dataset,
    )

    from backend.app.services.evaluation_pipeline_service import (
        process_and_publish_evaluation_collection,
    )

    assert_evaluation_database()

    documents = (
        resolve_evaluation_documents(
            xlsx_path=xlsx_path,
            sheet_name=sheet_name,
        )
    )

    print(
        "\n[Evaluation Dataset]"
    )

    print(
        f"dataset_id : "
        f"{dataset_id}"
    )

    print(
        f"문서 수    : "
        f"{len(documents)}"
    )

    for document in documents:

        print(
            "  - "
            f"{document.evaluation_document_id}"
            " | "
            f"{document.document_format}"
            " | "
            f"{document.source_path}"
        )

    registration = (
        register_evaluation_dataset(
            dataset_id=dataset_id,
            documents=[
                EvaluationDocumentInput(
                    evaluation_document_id=(
                        item.evaluation_document_id
                    ),
                    source_path=str(
                        item.source_path
                    ),
                    document_format=(
                        item.document_format
                    ),
                    title=item.title,
                )
                for item
                in documents
            ],
        )
    )

    collection_run_id = int(
        registration[
            "collection_run_id"
        ]
    )

    print(
        "\n[Evaluation DB 등록 완료]"
    )

    print(
        "collection_run_id : "
        f"{collection_run_id}"
    )

    print(
        "document_count    : "
        f"{registration.get('document_count')}"
    )

    print(
        "\n[실제 Document Pipeline 실행]"
    )

    pipeline_result = (
        process_and_publish_evaluation_collection(
            collection_run_id=(
                collection_run_id
            )
        )
    )

    document_map: dict[
        str,
        dict[str, Any],
    ] = {}

    for item in pipeline_result[
        "documents"
    ]:

        evaluation_document_id = str(
            item[
                "evaluation_document_id"
            ]
        ).strip()

        document_map[
            evaluation_document_id
        ] = {
            "announcement_id":
                item[
                    "announcement_id"
                ],

            "document_id":
                item[
                    "document_id"
                ],

            "processing_run_id":
                item[
                    "processing_run_id"
                ],

            "chunk_set_id":
                item[
                    "chunk_set_id"
                ],

            "chunk_count":
                item[
                    "chunk_count"
                ],

            "embedding_count":
                item[
                    "embedding_count"
                ],

            "embedding_model_name":
                item[
                    "embedding_model_name"
                ],
        }

    manifest = {
        "schema_version":
            "evaluation-pipeline-v1",

        "dataset_id":
            dataset_id,

        "dataset_xlsx":
            str(
                xlsx_path
            ),

        "collection_run_id":
            collection_run_id,

        "documents":
            document_map,

        "processing":
            pipeline_result.get(
                "processing"
            ),

        "publish":
            pipeline_result.get(
                "publish"
            ),
    }

    manifest_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with manifest_path.open(
        "w",
        encoding="utf-8",
    ) as file:

        json.dump(
            manifest,
            file,
            ensure_ascii=False,
            indent=2,
        )

        file.write(
            "\n"
        )

    print(
        "\n[Pipeline 완료]"
    )

    for (
        evaluation_document_id,
        information,
    ) in document_map.items():

        print(
            "  "
            f"{evaluation_document_id}"
            " -> "
            "announcement_id="
            f"{information['announcement_id']}"
            ", "
            "chunks="
            f"{information['chunk_count']}"
            ", "
            "embeddings="
            f"{information['embedding_count']}"
        )

    print(
        f"\nmanifest : "
        f"{manifest_path}"
    )

    return manifest


# ============================================================
# Manifest
# ============================================================


def load_manifest(
    *,
    manifest_path: Path,
    dataset_id: str,
) -> dict[str, Any]:

    if not manifest_path.is_file():

        raise FileNotFoundError(
            "평가 Pipeline manifest가 없습니다.\n"
            "--prepare를 먼저 실행하거나 "
            "--manifest 경로를 확인하세요.\n"
            f"{manifest_path}"
        )

    with manifest_path.open(
        "r",
        encoding="utf-8",
    ) as file:

        payload = json.load(
            file
        )

    if not isinstance(
        payload,
        dict,
    ):
        raise ValueError(
            "manifest 형식이 "
            "올바르지 않습니다."
        )

    if not isinstance(
        payload.get(
            "documents"
        ),
        dict,
    ):
        raise ValueError(
            "manifest의 documents가 "
            "올바르지 않습니다."
        )

    manifest_dataset = str(
        payload.get(
            "dataset_id"
        )
        or ""
    ).strip().upper()

    if (
        manifest_dataset
        and manifest_dataset
        != dataset_id.upper()
    ):
        raise ValueError(
            "현재 dataset과 manifest의 "
            "dataset이 다릅니다. "
            f"current={dataset_id}, "
            f"manifest={manifest_dataset}"
        )

    return payload


# ============================================================
# 실제 RAG 평가
# ============================================================


def evaluate(
    *,
    input_xlsx: Path,
    output_xlsx: Path,
    manifest: dict[str, Any],
    sheet_name: str,
    run_id: str,
    selected_question_ids:
        set[str] | None,
    rerun_success: bool,
    api_base_url: str,
    endpoint: str,
    timeout: int,
    retry_count: int,
) -> None:
    """
    문서 Pipeline 처리가 끝난 평가 DB를 대상으로
    실제 FastAPI /api/chat을 호출한다.

    /api/chat
        ↓
    Backend Chat Service
        ↓
    RAG Service
        ↓
    Hybrid Search
        ↓
    Vector + Keyword + RRF
        ↓
    Generation
        ↓
    answer + evidence
    """

    assert_evaluation_database()

    activation_result = (
        activate_evaluation_collection(
            manifest=manifest,
        )
    )

    print(
        "\n[Evaluation Collection 활성화]"
    )

    print(
        "previous : "
        f"{activation_result.get('previous_collection_run_id')}"
    )

    print(
        "active   : "
        f"{activation_result.get('active_collection_run_id')}"
    )

    print(
        "status   : "
        f"{activation_result.get('status')}"
    )

    if output_xlsx.is_file():

        wb = load_workbook(
            output_xlsx
        )

        resume_mode = True

    else:

        wb = load_workbook(
            input_xlsx
        )

        resume_mode = False

    try:
        if (
            sheet_name
            not in wb.sheetnames
        ):

            raise ValueError(
                f"'{sheet_name}' "
                "시트가 없습니다.\n"
                f"현재 시트: "
                f"{wb.sheetnames}"
            )

        ws = wb[
            sheet_name
        ]

        columns = (
            find_columns(
                ws
            )
        )

        required = [
            "document_id",
            "question_id",
            "user_input",
            "retrieved_chunk_ids",
            "retrieved_contexts",
            "response",
            "run_id",
            "git_commit",
        ]

        missing = [
            name
            for name
            in required
            if name
            not in columns
        ]

        if missing:

            raise ValueError(
                "평가셋에 필요한 "
                "column이 없습니다:\n"
                + ", ".join(
                    missing
                )
            )

        document_map = (
            manifest[
                "documents"
            ]
        )

        branch, git_commit = (
            get_git_info()
        )

        api_url = build_api_url(
            api_base_url=(
                api_base_url
            ),
            endpoint=endpoint,
        )

        processed = 0
        success = 0
        failed = 0
        skipped = 0

        found_question_ids: (
            set[str]
        ) = set()

        print(
            "=" * 78
        )

        print(
            "실제 RAG Pipeline 평가"
        )

        print(
            f"입력 파일       : "
            f"{input_xlsx}"
        )

        print(
            f"출력 파일       : "
            f"{output_xlsx}"
        )

        print(
            f"API             : "
            f"{api_url}"
        )

        print(
            "CollectionRun   : "
            f"{manifest.get('collection_run_id')}"
        )

        print(
            f"Resume          : "
            f"{resume_mode}"
        )

        print(
            f"run_id          : "
            f"{run_id}"
        )

        print(
            f"branch          : "
            f"{branch or '(확인 불가)'}"
        )

        print(
            f"git_commit      : "
            f"{git_commit or '(확인 불가)'}"
        )

        print(
            f"timeout         : "
            f"{timeout}s"
        )

        print(
            f"retry           : "
            f"{retry_count}"
        )

        print(
            "=" * 78
        )

        for row in range(
            2,
            ws.max_row + 1,
        ):

            # -----------------------------------------------
            # question_id
            # -----------------------------------------------

            question_id_raw = (
                ws.cell(
                    row=row,
                    column=columns[
                        "question_id"
                    ],
                ).value
            )

            question_id = (
                ""
                if question_id_raw
                is None
                else str(
                    question_id_raw
                ).strip().upper()
            )

            if (
                selected_question_ids
                is not None
                and question_id
                not in selected_question_ids
            ):
                continue

            if (
                selected_question_ids
                is not None
            ):
                found_question_ids.add(
                    question_id
                )

            # -----------------------------------------------
            # question
            # -----------------------------------------------

            question_raw = (
                ws.cell(
                    row=row,
                    column=columns[
                        "user_input"
                    ],
                ).value
            )

            if question_raw is None:
                continue

            question = str(
                question_raw
            ).strip()

            if not question:
                continue

            # -----------------------------------------------
            # document_id
            # -----------------------------------------------

            evaluation_document_id = str(
                ws.cell(
                    row=row,
                    column=columns[
                        "document_id"
                    ],
                ).value
                or ""
            ).strip()

            if (
                evaluation_document_id
                not in document_map
            ):

                raise RuntimeError(
                    "manifest에 "
                    "document_id가 없습니다: "
                    f"{evaluation_document_id}"
                )

            # -----------------------------------------------
            # Resume
            # -----------------------------------------------

            existing_response = str(
                ws.cell(
                    row=row,
                    column=columns[
                        "response"
                    ],
                ).value
                or ""
            ).strip()

            is_error_response = (
                existing_response.startswith(
                    "[API ERROR]"
                )
                or existing_response.startswith(
                    "[RAG ERROR]"
                )
            )

            if (
                existing_response
                and not is_error_response
                and not rerun_success
            ):

                skipped += 1

                print(
                    f"[RESUME] "
                    f"{question_id}: "
                    "기존 성공 결과 유지"
                )

                continue

            processed += 1

            document_info = (
                document_map[
                    evaluation_document_id
                ]
            )

            announcement_id = int(
                document_info[
                    "announcement_id"
                ]
            )

            embedding_model = str(
                document_info.get(
                    "embedding_model_name"
                )
                or ""
            )

            print(
                f"\n[{processed:02d}] "
                f"{question_id}: "
                f"{question}"
            )

            print(
                "  document_id     : "
                f"{evaluation_document_id}"
            )

            print(
                "  announcement_id : "
                f"{announcement_id}"
            )

            started = (
                time.perf_counter()
            )

            # -----------------------------------------------
            # /api/chat
            # -----------------------------------------------

            try:

                response_json = (
                    request_chat_with_retry(
                        api_url=api_url,
                        announcement_id=(
                            announcement_id
                        ),
                        question=question,
                        timeout=timeout,
                        retry_count=(
                            retry_count
                        ),
                    )
                )

            except Exception as exc:

                failed += 1

                # 이전 재실행 결과가 남는 것을 방지
                ws.cell(
                    row=row,
                    column=columns[
                        "retrieved_chunk_ids"
                    ],
                    value="",
                )

                ws.cell(
                    row=row,
                    column=columns[
                        "retrieved_contexts"
                    ],
                    value="",
                )

                ws.cell(
                    row=row,
                    column=columns[
                        "response"
                    ],
                    value=(
                        "[API ERROR] "
                        f"{type(exc).__name__}: "
                        f"{exc}"
                    ),
                )

                ws.cell(
                    row=row,
                    column=columns[
                        "run_id"
                    ],
                    value=run_id,
                )

                ws.cell(
                    row=row,
                    column=columns[
                        "git_commit"
                    ],
                    value=git_commit,
                )

                set_optional_cell(
                    ws=ws,
                    columns=columns,
                    row=row,
                    name="git_branch",
                    value=branch,
                )

                set_optional_cell(
                    ws=ws,
                    columns=columns,
                    row=row,
                    name="grounded",
                    value=None,
                )

                set_optional_cell(
                    ws=ws,
                    columns=columns,
                    row=row,
                    name="evidence",
                    value="",
                )

                set_optional_cell(
                    ws=ws,
                    columns=columns,
                    row=row,
                    name="embedding_model",
                    value=embedding_model,
                )

                elapsed_ms = int(
                    (
                        time.perf_counter()
                        - started
                    )
                    * 1000
                )

                set_optional_cell(
                    ws=ws,
                    columns=columns,
                    row=row,
                    name="elapsed_ms",
                    value=elapsed_ms,
                )

                output_xlsx.parent.mkdir(
                    parents=True,
                    exist_ok=True,
                )

                # 실패도 즉시 저장.
                # 다음 실행 시 error response는 다시 시도된다.
                wb.save(
                    output_xlsx
                )

                print(
                    "  실패             : "
                    f"{type(exc).__name__}: "
                    f"{exc}"
                )

                continue

            # -----------------------------------------------
            # Response 파싱
            # -----------------------------------------------

            answer = (
                response_json.get(
                    "answer"
                )
            )

            # 구버전 API 호환
            if answer is None:

                answer = (
                    response_json.get(
                        "response",
                        "",
                    )
                )

            answer = str(
                answer
                or ""
            )

            grounded = (
                response_json.get(
                    "grounded"
                )
            )

            evidence = (
                normalize_evidence(
                    response_json.get(
                        "evidence"
                    )
                )
            )

            chunk_ids = (
                serialize_chunk_ids(
                    evidence
                )
            )

            contexts = (
                serialize_contexts(
                    evidence
                )
            )

            elapsed_ms = int(
                (
                    time.perf_counter()
                    - started
                )
                * 1000
            )

            # -----------------------------------------------
            # 필수 결과 저장
            # -----------------------------------------------

            ws.cell(
                row=row,
                column=columns[
                    "retrieved_chunk_ids"
                ],
                value=chunk_ids,
            )

            ws.cell(
                row=row,
                column=columns[
                    "retrieved_contexts"
                ],
                value=contexts,
            )

            ws.cell(
                row=row,
                column=columns[
                    "response"
                ],
                value=answer,
            )

            ws.cell(
                row=row,
                column=columns[
                    "run_id"
                ],
                value=run_id,
            )

            ws.cell(
                row=row,
                column=columns[
                    "git_commit"
                ],
                value=git_commit,
            )

            # -----------------------------------------------
            # 데이터셋에 존재할 경우 부가정보 저장
            # -----------------------------------------------

            set_optional_cell(
                ws=ws,
                columns=columns,
                row=row,
                name="git_branch",
                value=branch,
            )

            set_optional_cell(
                ws=ws,
                columns=columns,
                row=row,
                name="grounded",
                value=grounded,
            )

            set_optional_cell(
                ws=ws,
                columns=columns,
                row=row,
                name="evidence",
                value=(
                    serialize_evidence_json(
                        evidence
                    )
                ),
            )

            set_optional_cell(
                ws=ws,
                columns=columns,
                row=row,
                name="embedding_model",
                value=embedding_model,
            )

            set_optional_cell(
                ws=ws,
                columns=columns,
                row=row,
                name="elapsed_ms",
                value=elapsed_ms,
            )

            output_xlsx.parent.mkdir(
                parents=True,
                exist_ok=True,
            )

            # -----------------------------------------------
            # 문항 1개마다 즉시 저장
            # -----------------------------------------------

            wb.save(
                output_xlsx
            )

            success += 1

            print(
                f"  answer           : "
                f"{answer[:100]}"
            )

            print(
                f"  grounded         : "
                f"{grounded}"
            )

            print(
                f"  evidence         : "
                f"{len(evidence)}개"
            )

            print(
                f"  chunk_ids        : "
                f"{chunk_ids or '(없음)'}"
            )

            print(
                f"  elapsed          : "
                f"{elapsed_ms / 1000:.2f}s"
            )

        # ====================================================
        # 지정 문항 존재 여부
        # ====================================================

        if (
            selected_question_ids
            is not None
        ):

            missing_question_ids = (
                selected_question_ids
                - found_question_ids
            )

            if missing_question_ids:

                print(
                    "\n[경고] Excel에서 "
                    "찾지 못한 문항: "
                    + ", ".join(
                        sorted(
                            missing_question_ids
                        )
                    )
                )

        print(
            "\n"
            + "=" * 78
        )

        print(
            "실제 Pipeline 평가 완료"
        )

        print(
            f"이번 처리  : "
            f"{processed}"
        )

        print(
            f"성공       : "
            f"{success}"
        )

        print(
            f"실패       : "
            f"{failed}"
        )

        print(
            f"Resume Skip: "
            f"{skipped}"
        )

        print(
            f"결과 파일  : "
            f"{output_xlsx}"
        )

        print(
            "=" * 78
        )

    finally:

        wb.close()


# ============================================================
# CLI
# ============================================================


def parse_args() -> argparse.Namespace:

    parser = (
        argparse.ArgumentParser(
            description=(
                "평가 원본 문서를 평가 DB에 등록하고 "
                "실제 Document Pipeline으로 처리한 뒤 "
                "FastAPI /api/chat을 호출하여 "
                "실제 서비스와 동일한 RAG 평가를 수행합니다."
            )
        )
    )

    parser.add_argument(
        "--dataset",
        default=None,
        help=(
            "평가셋 코드. "
            "GC/BD/DH/GP 등 새로운 코드도 "
            "소스 수정 없이 사용할 수 있습니다."
        ),
    )

    parser.add_argument(
        "--xlsx",
        default=None,
        help=(
            "평가셋 Excel 경로. "
            "생략하면 "
            "evaluation/datasets/"
            "<dataset>_FINAL_V*.xlsx 중 "
            "가장 높은 버전을 자동 선택합니다."
        ),
    )

    parser.add_argument(
        "--sheet",
        default=(
            DEFAULT_SHEET_NAME
        ),
    )

    # --------------------------------------------------------
    # Document Pipeline
    # --------------------------------------------------------

    parser.add_argument(
        "--prepare",
        action="store_true",
        help=(
            "평가 문서를 평가 DB에 등록하고 "
            "실제 Document Worker Pipeline으로 "
            "처리/Publish한 다음 "
            "질의응답 평가까지 이어서 실행합니다."
        ),
    )

    parser.add_argument(
        "--prepare-only",
        action="store_true",
        help=(
            "평가 DB 등록 + 실제 문서처리 + "
            "Publish까지만 실행하고 "
            "질의응답은 실행하지 않습니다."
        ),
    )

    parser.add_argument(
        "--manifest",
        default=None,
        help=(
            "평가 Pipeline manifest 경로. "
            "생략하면 "
            "evaluation/runtime/"
            "<dataset>_pipeline.json"
        ),
    )

    # --------------------------------------------------------
    # Run
    # --------------------------------------------------------

    parser.add_argument(
        "--run-number",
        default=(
            DEFAULT_RUN_NUMBER
        ),
        help=(
            "결과 파일 Run 번호. "
            "예: 001, 002, 003"
        ),
    )

    parser.add_argument(
        "--run-id",
        default=None,
        help=(
            "직접 run_id를 지정합니다. "
            "생략하면 "
            "<DATASET>_ACTUAL_RUN_<번호>"
        ),
    )

    parser.add_argument(
        "--output",
        default=None,
        help=(
            "결과 Excel 경로. "
            "생략하면 자동 생성합니다."
        ),
    )

    parser.add_argument(
        "--question-ids",
        default=None,
        help=(
            "특정 문항만 실행. "
            "예: Q001,Q002,Q010"
        ),
    )

    parser.add_argument(
        "--rerun-success",
        action="store_true",
        help=(
            "기존 결과 파일에 성공 response가 "
            "있어도 다시 실행합니다."
        ),
    )

    # --------------------------------------------------------
    # API
    # --------------------------------------------------------

    parser.add_argument(
        "--base-url",
        default=(
            DEFAULT_API_BASE_URL
        ),
        help=(
            "평가 Backend 주소. "
            "기본값: "
            "http://127.0.0.1:8000"
        ),
    )

    parser.add_argument(
        "--endpoint",
        default=(
            DEFAULT_ENDPOINT
        ),
        help=(
            "Chat endpoint. "
            "기본값: /api/chat"
        ),
    )

    parser.add_argument(
        "--timeout",
        type=int,
        default=(
            DEFAULT_TIMEOUT_SECONDS
        ),
        help=(
            "API 요청 timeout 초. "
            "기본 600초."
        ),
    )

    parser.add_argument(
        "--retry",
        type=int,
        default=(
            DEFAULT_RETRY_COUNT
        ),
        help=(
            "API 실패 시 추가 재시도 횟수. "
            "기본 1회."
        ),
    )

    args = (
        parser.parse_args()
    )

    if args.timeout <= 0:

        parser.error(
            "--timeout은 "
            "1 이상이어야 합니다."
        )

    if args.retry < 0:

        parser.error(
            "--retry는 "
            "0 이상이어야 합니다."
        )

    if (
        args.prepare_only
        and args.prepare
    ):

        parser.error(
            "--prepare와 "
            "--prepare-only는 "
            "동시에 지정하지 마세요."
        )

    return args


# ============================================================
# Main
# ============================================================


def main() -> None:

    args = (
        parse_args()
    )

    try:
        # ====================================================
        # Dataset 자동 탐색
        # ====================================================

        (
            dataset_id,
            input_xlsx,
            dataset_version,
        ) = resolve_dataset_xlsx(
            dataset=args.dataset,
            xlsx=args.xlsx,
        )

        # ====================================================
        # Manifest 경로
        # ====================================================

        manifest_path = (
            Path(
                args.manifest
            )
            if args.manifest
            else default_manifest_path(
                dataset_id=(
                    dataset_id
                )
            )
        )

        if not (
            manifest_path
            .is_absolute()
        ):

            manifest_path = (
                Path.cwd()
                / manifest_path
            ).resolve()

        # ====================================================
        # 실제 Document Pipeline
        # ====================================================

        if (
            args.prepare
            or args.prepare_only
        ):

            manifest = (
                prepare_actual_pipeline(
                    dataset_id=(
                        dataset_id
                    ),
                    xlsx_path=(
                        input_xlsx
                    ),
                    sheet_name=(
                        args.sheet
                    ),
                    manifest_path=(
                        manifest_path
                    ),
                )
            )

            print(
                "\n[PREPARE 완료] "
                "collection_run_id="
                f"{manifest['collection_run_id']}"
            )

            print(
                f"[MANIFEST] "
                f"{manifest_path}"
            )

            if (
                args.prepare_only
            ):
                return

        else:

            manifest = (
                load_manifest(
                    manifest_path=(
                        manifest_path
                    ),
                    dataset_id=(
                        dataset_id
                    ),
                )
            )

        # ====================================================
        # Output
        # ====================================================

        output_xlsx = (
            Path(
                args.output
            )
            if args.output
            else default_result_path(
                dataset_id=(
                    dataset_id
                ),
                dataset_version=(
                    dataset_version
                ),
                run_number=(
                    args.run_number
                ),
            )
        )

        if not (
            output_xlsx
            .is_absolute()
        ):

            output_xlsx = (
                Path.cwd()
                / output_xlsx
            ).resolve()

        # ====================================================
        # Run ID
        # ====================================================

        run_id = (
            args.run_id
            or (
                f"{dataset_id}_"
                f"ACTUAL_RUN_"
                f"{args.run_number}"
            )
        )

        # ====================================================
        # RAG 평가
        # ====================================================

        evaluate(
            input_xlsx=(
                input_xlsx
            ),
            output_xlsx=(
                output_xlsx
            ),
            manifest=manifest,
            sheet_name=(
                args.sheet
            ),
            run_id=(
                run_id
            ),
            selected_question_ids=(
                parse_question_ids(
                    args.question_ids
                )
            ),
            rerun_success=(
                args.rerun_success
            ),
            api_base_url=(
                args.base_url
            ),
            endpoint=(
                args.endpoint
            ),
            timeout=(
                args.timeout
            ),
            retry_count=(
                args.retry
            ),
        )

    except KeyboardInterrupt:

        print(
            "\n사용자가 평가를 중단했습니다.\n"
            "완료된 문항은 결과 Excel에 "
            "저장되어 있습니다."
        )

        sys.exit(
            130
        )

    except Exception as exc:

        print(
            "\n평가 실행 중 오류가 "
            "발생했습니다:\n"
            f"{type(exc).__name__}: "
            f"{exc}"
        )

        sys.exit(
            1
        )


if __name__ == "__main__":
    main()