from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from config.paths import PROJECT_ROOT
from pipeline.parser.format_detector import detect_actual_document_format


BASE_DIR = Path(__file__).resolve().parent
DATASETS_DIR = BASE_DIR / "datasets"
SOURCE_DOCUMENTS_DIR = BASE_DIR / "source_documents"
RESULTS_DIR = BASE_DIR / "results"
RUNTIME_DIR = BASE_DIR / "runtime"

DEFAULT_SHEET_NAME = "평가셋"

_DATASET_RE = re.compile(r"^[A-Za-z0-9_.-]+$")
_DATASET_FROM_NAME_RE = re.compile(
    r"^(?P<dataset>[A-Za-z0-9_.-]+?)_FINAL_V\d+",
    re.IGNORECASE,
)
_DATASET_VERSION_RE = re.compile(
    r"_FINAL_V(?P<version>\d+)",
    re.IGNORECASE,
)
_SOURCE_VERSION_RE = re.compile(
    r"^v(?P<version>\d+)$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class ResolvedEvaluationDocument:
    evaluation_document_id: str
    source_path: Path
    document_format: str
    title: str


def normalize_dataset_id(value: str) -> str:
    dataset_id = str(value or "").strip().upper()

    if not dataset_id:
        raise ValueError("dataset이 비어 있습니다.")

    if not _DATASET_RE.fullmatch(dataset_id):
        raise ValueError(
            "dataset에는 영문/숫자/._-만 사용할 수 있습니다: "
            f"{value}"
        )

    return dataset_id


def infer_dataset_id_from_path(path: Path) -> str:
    match = _DATASET_FROM_NAME_RE.match(path.stem)

    if match:
        return normalize_dataset_id(
            match.group("dataset")
        )

    prefix = path.stem.split("_", 1)[0]
    return normalize_dataset_id(prefix)


def dataset_version_from_path(path: Path) -> int:
    match = _DATASET_VERSION_RE.search(path.stem)

    if not match:
        return 1

    return int(match.group("version"))


def resolve_dataset_xlsx(
    *,
    dataset: str | None,
    xlsx: str | None,
) -> tuple[str, Path, int]:
    """
    평가셋 Excel을 찾는다.

    우선순위:
    1. --xlsx가 있으면 그 경로 사용
    2. --dataset만 있으면 evaluation/datasets/<CODE>_FINAL_V*.xlsx
       중 가장 높은 V 버전 사용

    GC/BD/DH/GP 같은 코드 목록을 소스에 하드코딩하지 않는다.
    """
    if xlsx:
        path = Path(xlsx).expanduser()

        if not path.is_absolute():
            path = PROJECT_ROOT / path

        path = path.resolve()

        if not path.is_file():
            raise FileNotFoundError(
                f"평가셋 Excel을 찾을 수 없습니다: {path}"
            )

        dataset_id = (
            normalize_dataset_id(dataset)
            if dataset
            else infer_dataset_id_from_path(path)
        )

        return (
            dataset_id,
            path,
            dataset_version_from_path(path),
        )

    if not dataset:
        raise ValueError(
            "--dataset 또는 --xlsx 중 하나는 필요합니다."
        )

    dataset_id = normalize_dataset_id(dataset)

    candidates = list(
        DATASETS_DIR.glob(
            f"{dataset_id}_FINAL_V*.xlsx"
        )
    )

    if not candidates:
        raise FileNotFoundError(
            "평가셋 Excel을 찾을 수 없습니다. "
            f"dataset={dataset_id}, dir={DATASETS_DIR}"
        )

    candidates.sort(
        key=lambda path: (
            dataset_version_from_path(path),
            path.name,
        ),
        reverse=True,
    )

    selected = candidates[0].resolve()

    return (
        dataset_id,
        selected,
        dataset_version_from_path(selected),
    )


def read_evaluation_document_ids(
    *,
    xlsx_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[str]:
    """
    평가셋의 document_id 열에서 사용되는 문서 ID를 중복 제거하여 읽는다.

    예:
      DOC_BD_001
      DOC_DH_001
    """
    wb = load_workbook(
        xlsx_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"'{sheet_name}' 시트가 없습니다. "
                f"현재 시트={wb.sheetnames}"
            )

        ws = wb[sheet_name]

        headers = {
            str(cell.value).strip(): cell.column
            for cell in ws[1]
            if cell.value is not None
        }

        if "document_id" not in headers:
            raise ValueError(
                "평가셋에 document_id 열이 없습니다."
            )

        document_ids: list[str] = []
        seen: set[str] = set()
        column = headers["document_id"]

        for row in range(2, ws.max_row + 1):
            raw = ws.cell(
                row=row,
                column=column,
            ).value

            if raw is None:
                continue

            document_id = str(raw).strip()

            if not document_id:
                continue

            if document_id in seen:
                continue

            seen.add(document_id)
            document_ids.append(document_id)

        if not document_ids:
            raise ValueError(
                "평가셋에서 document_id를 찾지 못했습니다."
            )

        return document_ids

    finally:
        wb.close()


def _source_version(path: Path) -> int:
    best = -1

    for part in path.parts:
        match = _SOURCE_VERSION_RE.fullmatch(part)

        if match:
            best = max(
                best,
                int(match.group("version")),
            )

    return best


def resolve_source_document(
    evaluation_document_id: str,
) -> ResolvedEvaluationDocument:
    """
    document_id와 동일한 폴더 아래에서 HWP/HWPX를 자동 탐색한다.

    예:
      evaluation/source_documents/DOC_BD_001/v1/공고문.hwpx

    v1, v2가 함께 있으면 가장 높은 버전을 선택한다.
    같은 최신 버전에 원본 문서가 여러 개면 모호하므로 실패시킨다.
    """
    document_dir = (
        SOURCE_DOCUMENTS_DIR
        / evaluation_document_id
    )

    if not document_dir.is_dir():
        raise FileNotFoundError(
            "평가 원본 문서 디렉터리를 찾을 수 없습니다: "
            f"{document_dir}"
        )

    candidates = [
        path
        for path in document_dir.rglob("*")
        if (
            path.is_file()
            and path.suffix.lower()
            in {".hwp", ".hwpx"}
        )
    ]

    if not candidates:
        raise FileNotFoundError(
            "평가 원본 HWP/HWPX가 없습니다: "
            f"{document_dir}"
        )

    candidates.sort(
        key=lambda path: (
            _source_version(path),
            path.name,
        ),
        reverse=True,
    )

    newest_version = _source_version(
        candidates[0]
    )

    newest_candidates = [
        path
        for path in candidates
        if _source_version(path)
        == newest_version
    ]

    if len(newest_candidates) != 1:
        raise RuntimeError(
            "같은 최신 버전에 평가 원본 문서가 여러 개 있습니다. "
            "버전 폴더를 정리하거나 원본문서를 하나로 유지해 주세요: "
            + ", ".join(
                str(path)
                for path in newest_candidates
            )
        )

    source_path = newest_candidates[0].resolve()

    actual_format = detect_actual_document_format(
        source_path
    )

    if actual_format not in {
        "hwp",
        "hwpx",
    }:
        raise RuntimeError(
            "실제 문서 형식을 HWP/HWPX로 판별하지 못했습니다: "
            f"{source_path}"
        )

    return ResolvedEvaluationDocument(
        evaluation_document_id=evaluation_document_id,
        source_path=source_path,
        document_format=actual_format,
        title=source_path.stem,
    )


def resolve_evaluation_documents(
    *,
    xlsx_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[ResolvedEvaluationDocument]:
    document_ids = read_evaluation_document_ids(
        xlsx_path=xlsx_path,
        sheet_name=sheet_name,
    )

    return [
        resolve_source_document(document_id)
        for document_id in document_ids
    ]


def default_manifest_path(
    *,
    dataset_id: str,
) -> Path:
    RUNTIME_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    return (
        RUNTIME_DIR
        / f"{dataset_id}_pipeline.json"
    )


def default_result_path(
    *,
    dataset_id: str,
    dataset_version: int,
    run_number: str,
) -> Path:
    RESULTS_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    return (
        RESULTS_DIR
        / (
            f"{dataset_id}_FINAL_V{dataset_version}_"
            f"ACTUAL_RUN_{run_number}_result.xlsx"
        )
    )


def resolve_result_xlsx(
    *,
    dataset: str | None,
    xlsx: str | None,
) -> tuple[str, Path]:
    """
    evaluate_metrics.py용 result.xlsx를 찾는다.

    --xlsx가 있으면 그대로 사용한다.
    --dataset만 있으면 ACTUAL_RUN 결과를 우선 탐색하고,
    없으면 일반 *_result.xlsx 중 가장 최근 수정 파일을 사용한다.
    """
    if xlsx:
        path = Path(xlsx).expanduser()

        if not path.is_absolute():
            path = PROJECT_ROOT / path

        path = path.resolve()

        if not path.is_file():
            raise FileNotFoundError(
                f"평가 결과 Excel을 찾을 수 없습니다: {path}"
            )

        dataset_id = (
            normalize_dataset_id(dataset)
            if dataset
            else infer_dataset_id_from_path(path)
        )

        return dataset_id, path

    if not dataset:
        raise ValueError(
            "--dataset 또는 --xlsx 중 하나는 필요합니다."
        )

    dataset_id = normalize_dataset_id(dataset)

    actual_candidates = list(
        RESULTS_DIR.glob(
            f"{dataset_id}_FINAL_V*_ACTUAL_RUN_*_result.xlsx"
        )
    )

    candidates = actual_candidates or list(
        RESULTS_DIR.glob(
            f"{dataset_id}_FINAL_V*_result.xlsx"
        )
    )

    if not candidates:
        raise FileNotFoundError(
            "평가 결과 Excel을 찾을 수 없습니다. "
            "먼저 evaluate_rag.py를 실행하거나 --xlsx를 지정하세요. "
            f"dataset={dataset_id}"
        )

    selected = max(
        candidates,
        key=lambda path: path.stat().st_mtime,
    )

    return dataset_id, selected.resolve()


def default_scored_path(
    result_xlsx: Path,
) -> Path:
    name = result_xlsx.name

    if name.endswith("_result.xlsx"):
        name = (
            name[:-len("_result.xlsx")]
            + "_scored.xlsx"
        )
    else:
        name = (
            result_xlsx.stem
            + "_scored.xlsx"
        )

    return result_xlsx.with_name(name)
