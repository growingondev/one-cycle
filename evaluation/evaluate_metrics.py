from __future__ import annotations

import argparse
import asyncio
import math
import os
import re
import sys
import time
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# ============================================================
# 기본 경로 / 설정
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "results"

DEFAULT_SHEET_NAME = "평가셋"

DEFAULT_RAGAS_BASE_URL = os.getenv(
    "RAGAS_API_BASE_URL",
    "http://127.0.0.1:8080/v1",
)

DEFAULT_RAGAS_API_KEY = os.getenv(
    "RAGAS_API_KEY",
    "no-key",
)

DEFAULT_RAGAS_MODEL = os.getenv(
    "RAGAS_MODEL",
    "",
).strip()



# ============================================================
# Recall 판정 기준
# ============================================================

NUMERIC_COVERAGE_THRESHOLD = 0.70
TOKEN_COVERAGE_THRESHOLD = 0.45

NO_NUMBER_TOKEN_THRESHOLD = 0.55
NO_NUMBER_SIMILARITY_THRESHOLD = 0.35

HIGH_SIMILARITY_THRESHOLD = 0.75


# ============================================================
# Dataset
# ============================================================


def normalize_dataset_name(
    value: str,
) -> str:
    dataset = value.strip().upper()

    aliases = {
        "GC": "GC",
        "GOCHANG": "GC",
        "고창": "GC",

        "BD": "BD",
        "BUNDONG": "BD",
        "서울번동": "BD",
    }

    if dataset not in aliases:
        raise ValueError(
            f"지원하지 않는 dataset입니다: {value}\n"
            "사용 가능: GC, BD"
        )

    return aliases[dataset]


def dataset_paths(
    dataset: str,
) -> tuple[Path, Path]:
    dataset = normalize_dataset_name(
        dataset
    )

    input_xlsx = (
        RESULTS_DIR
        / f"{dataset}_FINAL_V1_result.xlsx"
    )

    output_xlsx = (
        RESULTS_DIR
        / f"{dataset}_FINAL_V1_scored.xlsx"
    )

    return (
        input_xlsx,
        output_xlsx,
    )


# ============================================================
# Excel
# ============================================================


def find_columns(
    ws,
) -> dict[str, int]:
    columns: dict[str, int] = {}

    for cell in ws[1]:
        if cell.value is not None:
            columns[
                str(cell.value).strip()
            ] = cell.column

    return columns


def ensure_column(
    ws,
    columns: dict[str, int],
    name: str,
) -> int:
    if name in columns:
        return columns[name]

    new_column = ws.max_column + 1

    ws.cell(
        row=1,
        column=new_column,
        value=name,
    )

    columns[name] = new_column

    return new_column


# ============================================================
# retrieved_contexts 분리
# ============================================================


def split_retrieved_contexts(
    value: Any,
) -> list[str]:
    """
    evaluate_rag.py에서 저장한 retrieved_contexts를
    Rank별 Context 리스트로 변환한다.
    """

    if value is None:
        return []

    text = str(value).strip()

    if not text:
        return []

    parts = re.split(
        r"\n\s*---\s*\n",
        text,
    )

    contexts: list[str] = []

    for part in parts:
        part = part.strip()

        if not part:
            continue

        lines = part.splitlines()

        if (
            lines
            and lines[0]
            .strip()
            .startswith("[rank=")
        ):
            part = "\n".join(
                lines[1:]
            ).strip()

        if part:
            contexts.append(
                part
            )

    return contexts


# ============================================================
# Recall용 텍스트 정규화
# ============================================================


def normalize_text(
    text: str,
) -> str:
    """
    Recall 비교용 정규화.
    """

    text = str(text).lower()

    # 2026 → 26
    text = re.sub(
        r"\b20(\d{2})\b",
        r"\1",
        text,
    )

    # 10시 → 10:00
    text = re.sub(
        r"(\d{1,2})\s*시",
        r"\1:00",
        text,
    )

    # 공백 제거
    text = re.sub(
        r"\s+",
        "",
        text,
    )

    # 한글 / 영어 / 숫자만 유지
    text = re.sub(
        r"[^0-9a-z가-힣]",
        "",
        text,
    )

    return text


# ============================================================
# 숫자 정규화
# ============================================================


def normalize_number(
    value: str,
) -> str:
    """
    숫자 표기 통일.

    08 → 8
    08.0 → 8
    """

    value = value.strip()

    try:
        number = float(
            value
        )

        if number.is_integer():
            return str(
                int(number)
            )

        return str(
            number
        )

    except ValueError:
        return value


def extract_numbers(
    text: str,
) -> list[str]:
    """
    날짜 / 시간 / 금액 / 면적 / 나이 등의
    숫자 정보를 추출한다.
    """

    text = str(
        text
    )

    # 2026 → 26
    text = re.sub(
        r"\b20(\d{2})\b",
        r"\1",
        text,
    )

    numbers = re.findall(
        r"\d+(?:\.\d+)?",
        text,
    )

    return [
        normalize_number(
            number
        )
        for number
        in numbers
    ]


# ============================================================
# 토큰 추출
# ============================================================


def extract_tokens(
    text: str,
) -> list[str]:
    """
    reference_text에서 의미 비교에 사용할
    핵심 토큰을 추출한다.
    """

    text = str(
        text
    ).lower()

    text = re.sub(
        r"\b20(\d{2})\b",
        r"\1",
        text,
    )

    tokens = re.findall(
        r"[가-힣]{2,}"
        r"|[A-Za-z]+"
        r"|\d+(?:\.\d+)?",
        text,
    )

    stopwords = {
        "및",
        "또는",
        "그리고",
        "경우",
        "현재",
        "기준",
        "관련",
        "대한",
        "통해",
        "있으며",
        "있음",
        "해당",
        "한다",
        "됩니다",
        "한다면",
    }

    result: list[str] = []

    for token in tokens:
        token = (
            token
            .strip()
            .lower()
        )

        if not token:
            continue

        if token in stopwords:
            continue

        result.append(
            token
        )

    return result


# ============================================================
# Token Coverage
# ============================================================


def token_coverage(
    reference_text: str,
    context: str,
) -> float:
    """
    reference_text의 핵심 토큰 중
    context에 포함된 비율을 계산한다.
    """

    tokens = extract_tokens(
        reference_text
    )

    if not tokens:
        return 0.0

    context_norm = normalize_text(
        context
    )

    matched = 0

    for token in tokens:
        token_norm = normalize_text(
            token
        )

        if (
            token_norm
            and token_norm
            in context_norm
        ):
            matched += 1

    return (
        matched
        / len(tokens)
    )


# ============================================================
# Numeric Coverage
# ============================================================


def number_coverage(
    reference_text: str,
    context: str,
) -> float | None:
    """
    reference_text의 숫자 정보가
    context 안에 얼마나 존재하는지 계산한다.

    reference_text에 숫자가 없다면 None.
    """

    reference_numbers = (
        extract_numbers(
            reference_text
        )
    )

    if not reference_numbers:
        return None

    context_numbers = (
        extract_numbers(
            context
        )
    )

    if not context_numbers:
        return 0.0

    context_number_set = set(
        context_numbers
    )

    matched = sum(
        1
        for number
        in reference_numbers
        if number
        in context_number_set
    )

    return (
        matched
        / len(reference_numbers)
    )


# ============================================================
# 부분 문자열 유사도
# ============================================================


def partial_similarity(
    reference_text: str,
    context: str,
) -> float:
    """
    긴 Context의 일부 구간이
    reference_text와 얼마나 유사한지 계산한다.
    """

    ref_norm = normalize_text(
        reference_text
    )

    ctx_norm = normalize_text(
        context
    )

    if (
        not ref_norm
        or not ctx_norm
    ):
        return 0.0

    if ref_norm in ctx_norm:
        return 1.0

    if (
        len(ctx_norm)
        <= len(ref_norm)
    ):
        return SequenceMatcher(
            None,
            ref_norm,
            ctx_norm,
        ).ratio()

    window_size = max(
        len(ref_norm),
        int(
            len(ref_norm)
            * 2.0
        ),
    )

    step = max(
        1,
        len(ref_norm) // 5,
    )

    best_score = 0.0

    max_start = max(
        1,
        len(ctx_norm)
        - window_size
        + 1,
    )

    for start in range(
        0,
        max_start,
        step,
    ):
        candidate = ctx_norm[
            start:
            start + window_size
        ]

        score = SequenceMatcher(
            None,
            ref_norm,
            candidate,
        ).ratio()

        best_score = max(
            best_score,
            score,
        )

    if (
        len(ctx_norm)
        > window_size
    ):
        candidate = ctx_norm[
            -window_size:
        ]

        score = SequenceMatcher(
            None,
            ref_norm,
            candidate,
        ).ratio()

        best_score = max(
            best_score,
            score,
        )

    return best_score


# ============================================================
# Evidence 판정
# ============================================================


def evidence_matches(
    reference_text: str,
    context: str,
) -> tuple[
    bool,
    float,
    str,
]:
    """
    Context 안에 reference_text의 정답 근거가
    존재하는지 판단한다.

    판정 순서
    --------
    1. 정규화 후 완전 포함
    2. 숫자 + 핵심 토큰 기반 사실 일치
    3. 숫자가 없는 문장의 의미 토큰 일치
    4. 높은 문자열 유사도
    5. 복합 Evidence 보완 판정
    """

    ref_norm = normalize_text(
        reference_text
    )

    ctx_norm = normalize_text(
        context
    )

    if (
        not ref_norm
        or not ctx_norm
    ):
        return (
            False,
            0.0,
            "empty",
        )

    # ========================================================
    # 1. 정규화 후 완전 포함
    # ========================================================

    if ref_norm in ctx_norm:
        return (
            True,
            1.0,
            "normalized_exact_containment",
        )

    # ========================================================
    # 개별 점수 계산
    # ========================================================

    word_score = token_coverage(
        reference_text,
        context,
    )

    numeric_score = number_coverage(
        reference_text,
        context,
    )

    similarity = partial_similarity(
        reference_text,
        context,
    )

    # ========================================================
    # 2. 숫자 정보가 있는 정답
    # ========================================================

    if numeric_score is not None:

        # 숫자 대부분 + 핵심 단어 충분히 일치
        if (
            numeric_score
            >= NUMERIC_COVERAGE_THRESHOLD
            and word_score
            >= TOKEN_COVERAGE_THRESHOLD
        ):
            score = (
                0.55 * numeric_score
                + 0.35 * word_score
                + 0.10 * similarity
            )

            return (
                True,
                score,
                "numeric_fact_match",
            )

        # 숫자가 거의 전부 일치
        if (
            numeric_score >= 0.95
            and similarity >= 0.35
        ):
            score = (
                0.60 * numeric_score
                + 0.40 * similarity
            )

            return (
                True,
                score,
                "numeric_full_match",
            )

        # ----------------------------------------------------
        # 숫자 근거 보완 판정
        #
        # 날짜/시간/금액 등 핵심 숫자가 매우 잘 맞고,
        # 텍스트 쪽 근거도 조금이라도 있어야 HIT
        # ----------------------------------------------------

        if (
            numeric_score >= 0.80
            and (
                word_score >= 0.25
                or similarity >= 0.25
            )
        ):
            score = (
                0.60 * numeric_score
                + 0.25 * word_score
                + 0.15 * similarity
            )

            return (
                True,
                score,
                "numeric_evidence_match",
            )

    # ========================================================
    # 3. 숫자가 없는 일반 문장
    # ========================================================

    if numeric_score is None:

        if (
            word_score
            >= NO_NUMBER_TOKEN_THRESHOLD
            and similarity
            >= NO_NUMBER_SIMILARITY_THRESHOLD
        ):
            score = (
                0.70 * word_score
                + 0.30 * similarity
            )

            return (
                True,
                score,
                "semantic_token_match",
            )

        if (
            word_score >= 0.70
            and similarity >= 0.25
        ):
            score = (
                0.75 * word_score
                + 0.25 * similarity
            )

            return (
                True,
                score,
                "strong_token_match",
            )

    # ========================================================
    # 4. 문자열 자체가 매우 유사
    # ========================================================

    if (
        similarity
        >= HIGH_SIMILARITY_THRESHOLD
    ):
        return (
            True,
            similarity,
            "high_text_similarity",
        )

    # ========================================================
    # 5. 복합 Evidence 보완 판정
    #
    # 숫자 하나만 우연히 겹치는 것을 막기 위해
    # 서로 다른 Evidence가 최소 2개 이상
    # 일정 수준 이상이어야 한다.
    # ========================================================

    evidence_count = 0

    if word_score >= 0.50:
        evidence_count += 1

    if similarity >= 0.40:
        evidence_count += 1

    if (
        numeric_score is not None
        and numeric_score >= 0.70
    ):
        evidence_count += 1

    if evidence_count >= 2:

        if numeric_score is None:
            final_score = (
                0.65 * word_score
                + 0.35 * similarity
            )

        else:
            final_score = (
                0.45 * numeric_score
                + 0.35 * word_score
                + 0.20 * similarity
            )

        if final_score >= 0.55:
            return (
                True,
                final_score,
                "multi_evidence_match",
            )

    # ========================================================
    # 6. 실패
    # ========================================================

    candidates = [
        word_score,
        similarity,
    ]

    if numeric_score is not None:
        candidates.append(
            numeric_score
        )

    diagnostic_score = max(
        candidates
    )

    return (
        False,
        diagnostic_score,
        "no_match",
    )


# ============================================================
# Recall@K
# ============================================================


def recall_at_k(
    reference_text: str,
    contexts: list[str],
    k: int,
) -> tuple[
    int | None,
    int | None,
    float,
    str,
]:
    """
    Recall@K 계산.

    1. Top-K의 Context를 개별적으로 검사
    2. 개별 Context에서 실패하면
       Top-K 전체를 합쳐서 다시 검사

    비교형 / 복합형 질문에서
    정답 근거가 여러 청크에 나뉜 경우 대응.
    """

    if not reference_text.strip():
        return (
            None,
            None,
            0.0,
            "no_reference_text",
        )

    top_k_contexts = contexts[:k]

    if not top_k_contexts:
        return (
            0,
            None,
            0.0,
            "no_contexts",
        )

    best_score = 0.0
    best_reason = "no_match"

    # ========================================================
    # 1. 개별 Context 검사
    # ========================================================

    for rank, context in enumerate(
        top_k_contexts,
        start=1,
    ):
        (
            matched,
            score,
            reason,
        ) = evidence_matches(
            reference_text,
            context,
        )

        if score > best_score:
            best_score = score
            best_reason = reason

        if matched:
            return (
                1,
                rank,
                score,
                reason,
            )

    # ========================================================
    # 2. Top-K Context 통합 검사
    # ========================================================

    combined_context = "\n".join(
        top_k_contexts
    )

    (
        combined_matched,
        combined_score,
        combined_reason,
    ) = evidence_matches(
        reference_text,
        combined_context,
    )

    if combined_score > best_score:
        best_score = combined_score
        best_reason = combined_reason

    if combined_matched:
        return (
            1,
            None,
            combined_score,
            (
                f"combined_top_{k}_"
                f"{combined_reason}"
            ),
        )

    # ========================================================
    # 3. Top-K 전체에서도 실패
    # ========================================================

    return (
        0,
        None,
        best_score,
        best_reason,
    )


# ============================================================
# 선택 평가 문항 파싱
# ============================================================


def parse_question_ids(
    value: str | None,
) -> set[str] | None:
    """
    --question-ids로 전달된 문항 ID를 set으로 변환한다.

    예:
    Q002,Q003,Q007
        ↓
    {"Q002", "Q003", "Q007"}

    옵션을 사용하지 않으면 None을 반환하여
    기존처럼 전체 문항을 평가한다.
    """

    if value is None:
        return None

    question_ids = {
        item.strip().upper()
        for item in value.split(",")
        if item.strip()
    }

    return (
        question_ids
        or None
    )


# ============================================================
# RAGAS 패키지 확인
# ============================================================


def check_ragas_packages() -> None:
    missing: list[str] = []

    try:
        import ragas  # noqa: F401
    except ImportError:
        missing.append(
            "ragas"
        )

    try:
        import openai  # noqa: F401
    except ImportError:
        missing.append(
            "openai"
        )


    if missing:
        raise RuntimeError(
            "RAGAS 평가에 필요한 "
            "패키지가 없습니다.\n\n"
            "다음 명령으로 설치하세요:\n\n"
            f"pip install {' '.join(missing)}"
        )


# ============================================================
# RAGAS Scorer 생성
# ============================================================


async def build_ragas_scorers(
    base_url: str,
    api_key: str,
    model: str,
    adapt_factual_korean: bool = False,
    factual_only: bool = False,
) -> dict[str, Any]:

    from openai import AsyncOpenAI

    from ragas.llms import (
        llm_factory,
    )

    from ragas.metrics.collections import (
        Faithfulness,
        FactualCorrectness,
    )

    if not model:
        raise RuntimeError(
            "RAGAS 평가용 모델명이 없습니다.\n\n"
            "예:\n"
            "python evaluation/evaluate_metrics.py "
            "--dataset GC "
            "--ragas-model 모델명"
        )

    client = AsyncOpenAI(
        api_key=api_key,
        base_url=base_url,
    )

    llm = llm_factory(
        model,
        client=client,
        max_tokens=4096,
    )

    # ========================================================
    # Factual Correctness
    # ========================================================

    factual_correctness = (
        FactualCorrectness(
            llm=llm
        )
    )

    if adapt_factual_korean:

        print(
            "[RAGAS] FactualCorrectness "
            "한국어 Prompt adaptation 시작"
        )

        factual_correctness.prompt = (
            await factual_correctness.prompt.adapt(
                target_language="korean",
                llm=llm,
                adapt_instruction=True,
            )
        )

        factual_correctness.nli_prompt = (
            await factual_correctness.nli_prompt.adapt(
                target_language="korean",
                llm=llm,
                adapt_instruction=True,
            )
        )

        print(
            "[RAGAS] FactualCorrectness "
            "한국어 Prompt adaptation 완료"
        )

        print(
            "  Claim Prompt Language : "
            f"{factual_correctness.prompt.language}"
        )

        print(
            "  NLI Prompt Language   : "
            f"{factual_correctness.nli_prompt.language}"
        )

    scorers = {
        "factual_correctness":
            factual_correctness,
    }

    if not factual_only:
        scorers[
            "faithfulness"
        ] = Faithfulness(
            llm=llm
        )

    return scorers


# ============================================================
# RAGAS 문항 1개 평가
# ============================================================


async def score_one_with_ragas(
    scorers: dict[str, Any],
    user_input: str,
    reference: str,
    response: str,
    contexts: list[str],
    factual_only: bool = False,
    run_faithfulness: bool = True,
    run_factual_correctness: bool = True,
) -> tuple[
    dict[str, float | None],
    dict[str, float],
]:

    scores: dict[
        str,
        float | None,
    ] = {
        "faithfulness": None,
        "factual_correctness": None,
    }

    metric_times: dict[str, float] = {}

    async def safe_score(
        name: str,
        **kwargs,
    ) -> float | None:

        metric_start = time.perf_counter()

        try:
            result = await scorers[
                name
            ].ascore(
                **kwargs
            )

            value = result.value

            if value is None:
                return None

            value = float(value)

            if math.isnan(value):
                return None

            return value

        except Exception as exc:
            print(
                "    [RAGAS 경고] "
                f"{name} 실패: {exc}"
            )
            return None

        finally:
            elapsed = (
                time.perf_counter()
                - metric_start
            )

            metric_times[name] = elapsed

            print(
                "  [TIME] "
                f"{name:22s}: "
                f"{elapsed:8.2f}s"
            )

    if (
        not factual_only
        and run_faithfulness
    ):
        scores["faithfulness"] = await safe_score(
            "faithfulness",
            user_input=user_input,
            response=response,
            retrieved_contexts=contexts,
        )

    if run_factual_correctness:
        scores[
            "factual_correctness"
        ] = await safe_score(
            "factual_correctness",
            response=response,
            reference=reference,
        )

    return scores, metric_times


# ============================================================
# UNANSWERABLE / Correct Rejection
# ============================================================

REFUSAL_PATTERNS = (
    "제공된 LH 공고문에서 확인할 수 없습니다.",
    "확인할 수 없습니다",
    "확인되지 않습니다",
    "찾을 수 없습니다",
    "문서에서 확인할 수 없습니다",
    "문서에서 확인되지 않습니다",
    "공고문에서 확인할 수 없습니다",
    "공고문에서 확인되지 않습니다",
    "해당 정보가 없습니다",
    "관련 정보가 없습니다",
    "근거를 찾을 수 없습니다",
    "답변할 수 없습니다",
    "알 수 없습니다",
)


def score_correct_rejection(
    expected_behavior: str,
    response: str,
) -> tuple[int | None, str]:

    behavior = str(
        expected_behavior
        or ""
    ).strip().lower()

    if behavior not in {
        "refuse",
        "unanswerable",
    }:
        return (
            None,
            "N/A - answerable",
        )

    normalized = re.sub(
        r"\s+",
        " ",
        str(
            response
            or ""
        ).strip(),
    )

    if not normalized:
        return (
            0,
            "empty_response",
        )

    for pattern in REFUSAL_PATTERNS:
        if pattern in normalized:
            return (
                1,
                f"refusal_pattern={pattern}",
            )

    return (
        0,
        "no_refusal_pattern",
    )


# ============================================================
# 전체 평가
# ============================================================


async def evaluate_metrics(
    args: argparse.Namespace,
) -> None:

    dataset = normalize_dataset_name(
        args.dataset
    )

    (
        default_input,
        default_output,
    ) = dataset_paths(
        dataset
    )

    input_path = (
        Path(args.xlsx)
        if args.xlsx
        else default_input
    )

    output_path = (
        Path(args.output)
        if args.output
        else default_output
    )

    # ========================================================
    # 입력 파일 확인
    # ========================================================

    if not input_path.exists():
        raise FileNotFoundError(
            "평가 결과 파일을 "
            "찾을 수 없습니다:\n"
            f"{input_path}\n\n"
            "먼저 evaluate_rag.py를 "
            "실행하세요."
        )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    # ========================================================
    # Excel 읽기 / 자동 Resume
    # ========================================================

    resume_source = input_path

    if (
        output_path.exists()
        and not args.rerun_success
    ):
        resume_source = output_path
        print(
            "[RESUME] 기존 출력 파일에서 이어서 평가합니다."
        )
        print(
            f"[RESUME] 파일: {resume_source}"
        )

    wb = load_workbook(
        resume_source
    )

    if (
        args.sheet
        not in wb.sheetnames
    ):
        raise ValueError(
            f"'{args.sheet}' "
            "시트가 없습니다.\n"
            f"현재 시트: "
            f"{wb.sheetnames}"
        )

    ws = wb[
        args.sheet
    ]

    columns = find_columns(
        ws
    )

    # ========================================================
    # 선택 평가 문항
    # ========================================================

    selected_question_ids = (
        parse_question_ids(
            args.question_ids
        )
    )

    found_question_ids: set[str] = set()

    # ========================================================
    # 필수 열 확인
    # ========================================================

    required = [
        "question_id",
        "user_input",
        "reference",
        "reference_text",
        "retrieved_contexts",
        "response",
        "expected_behavior",
        "recall_at_1",
        "recall_at_3",
        "recall_at_5",
        "faithfulness",
        "factual_correctness",
    ]

    missing = [
        name
        for name in required
        if name not in columns
    ]

    if missing:
        raise ValueError(
            "평가에 필요한 Excel 열이 "
            "없습니다:\n"
            + ", ".join(
                missing
            )
        )

    # ========================================================
    # Recall 진단용 열
    # ========================================================

    recall_method_col = ensure_column(
        ws,
        columns,
        "recall_match_method",
    )

    recall_rank_col = ensure_column(
        ws,
        columns,
        "recall_matched_rank",
    )

    recall_score_col = ensure_column(
        ws,
        columns,
        "recall_match_score",
    )

    ragas_status_col = ensure_column(
        ws,
        columns,
        "ragas_status",
    )

    # 기존 결과 파일 호환을 위해 response_relevancy 열은 유지한다.
    # 새 RUN에서는 계산하지 않고 빈칸으로 둔다.
    response_relevancy_col = ensure_column(
        ws,
        columns,
        "response_relevancy",
    )

    correct_rejection_col = ensure_column(
        ws,
        columns,
        "correct_rejection",
    )

    rejection_reason_col = ensure_column(
        ws,
        columns,
        "rejection_match_reason",
    )

    # ========================================================
    # RAGAS 준비
    # ========================================================

    scorers = None

    if not args.skip_ragas:

        check_ragas_packages()

        scorers = (
            await build_ragas_scorers(
                base_url=(
                    args.ragas_base_url
                ),
                api_key=(
                    args.ragas_api_key
                ),
                model=(
                    args.ragas_model
                ),
                adapt_factual_korean=(
                    args.adapt_factual_korean
                ),
                factual_only=(
                    args.factual_only
                ),
            )
        )

    # ========================================================
    # 시작 정보
    # ========================================================

    print(
        "=" * 78
    )

    print(
        "RAG 평가 지표 계산"
    )

    print(
        f"dataset          : "
        f"{dataset}"
    )

    print(
        f"입력 파일       : "
        f"{input_path}"
    )

    print(
        f"출력 파일       : "
        f"{output_path}"
    )

    print(
        "Resume           : "
        + (
            "기존 성공 metric 자동 SKIP"
            if not args.rerun_success
            else "성공 metric도 강제 재실행"
        )
    )

    print(
        "Recall 방식      : "
        "Hybrid Evidence Matching "
        "+ Combined Top-K Context"
    )

    print(
        f"RAGAS            : "
        f"{'실행 안 함' if args.skip_ragas else '실행'}"
    )

    print(
        "평가 모드       : "
        + (
            "Factual Correctness only"
            if args.factual_only
            else "RAGAS 핵심 2개 metrics"
        )
    )

    print(
        "Factual Prompt   : "
        + (
            "한국어 Adaptation"
            if args.adapt_factual_korean
            else "기본 Prompt"
        )
    )

    if selected_question_ids:
        print(
            "선택 문항       : "
            + ", ".join(
                sorted(
                    selected_question_ids
                )
            )
        )
    else:
        print(
            "선택 문항       : 전체"
        )

    print(
        "=" * 78
    )

    # ========================================================
    # 통계 변수
    # ========================================================

    processed = 0
    answerable = 0
    unanswerable = 0
    correct_rejection_hits = 0

    recall_1_hits = 0
    recall_3_hits = 0
    recall_5_hits = 0

    ragas_values: dict[
        str,
        list[float],
    ] = {
        "faithfulness": [],
        "factual_correctness": [],
    }

    metric_time_totals: dict[
        str,
        float,
    ] = {
        "recall": 0.0,
        "faithfulness": 0.0,
        "factual_correctness": 0.0,
    }

    evaluation_start = time.perf_counter()

    # ========================================================
    # 문항 반복
    # ========================================================

    for row in range(
        2,
        ws.max_row + 1,
    ):

        question_id = ws.cell(
            row=row,
            column=columns[
                "question_id"
            ],
        ).value

        question_id = (
            ""
            if question_id is None
            else str(
                question_id
            ).strip().upper()
        )

        if (
            selected_question_ids is not None
            and question_id
            not in selected_question_ids
        ):
            continue

        if selected_question_ids is not None:
            found_question_ids.add(
                question_id
            )

        user_input = ws.cell(
            row=row,
            column=columns[
                "user_input"
            ],
        ).value

        if user_input is None:
            continue

        user_input = (
            str(user_input)
            .strip()
        )

        if not user_input:
            continue

        reference = ws.cell(
            row=row,
            column=columns[
                "reference"
            ],
        ).value

        reference_text = ws.cell(
            row=row,
            column=columns[
                "reference_text"
            ],
        ).value

        retrieved_raw = ws.cell(
            row=row,
            column=columns[
                "retrieved_contexts"
            ],
        ).value

        response = ws.cell(
            row=row,
            column=columns[
                "response"
            ],
        ).value

        reference = (
            ""
            if reference is None
            else str(
                reference
            ).strip()
        )

        reference_text = (
            ""
            if reference_text is None
            else str(
                reference_text
            ).strip()
        )

        response = (
            ""
            if response is None
            else str(
                response
            ).strip()
        )

        expected_behavior = str(
            ws.cell(
                row=row,
                column=columns[
                    "expected_behavior"
                ],
            ).value
            or ""
        ).strip().lower()

        existing_faithfulness = ws.cell(
            row=row,
            column=columns[
                "faithfulness"
            ],
        ).value

        existing_factual = ws.cell(
            row=row,
            column=columns[
                "factual_correctness"
            ],
        ).value

        has_faithfulness = (
            existing_faithfulness is not None
            and str(existing_faithfulness).strip() != ""
        )

        has_factual = (
            existing_factual is not None
            and str(existing_factual).strip() != ""
        )

        # 최종 성능에서 제외된 Response Relevancy는
        # 기존 열만 유지하고 새 RUN에서는 빈칸으로 둔다.
        ws.cell(
            row=row,
            column=response_relevancy_col,
            value=None,
        )

        (
            correct_rejection,
            rejection_reason,
        ) = score_correct_rejection(
            expected_behavior,
            response,
        )

        ws.cell(
            row=row,
            column=correct_rejection_col,
            value=correct_rejection,
        )

        ws.cell(
            row=row,
            column=rejection_reason_col,
            value=rejection_reason,
        )

        if correct_rejection is not None:
            unanswerable += 1
            correct_rejection_hits += (
                correct_rejection
            )

        contexts = (
            split_retrieved_contexts(
                retrieved_raw
            )
        )

        processed += 1

        question_start = time.perf_counter()

        print(
            f"\n[{processed:02d}] "
            f"{question_id}: "
            f"{user_input}"
        )

        # ====================================================
        # Recall@1 / @3 / @5 실행시간 측정
        # ====================================================

        recall_start = time.perf_counter()

        (
            r1,
            rank1,
            score1,
            reason1,
        ) = recall_at_k(
            reference_text,
            contexts,
            1,
        )

        # ====================================================
        # Recall@3
        # ====================================================

        (
            r3,
            rank3,
            score3,
            reason3,
        ) = recall_at_k(
            reference_text,
            contexts,
            3,
        )

        # ====================================================
        # Recall@5
        # ====================================================

        (
            r5,
            rank5,
            score5,
            reason5,
        ) = recall_at_k(
            reference_text,
            contexts,
            5,
        )

        recall_elapsed = (
            time.perf_counter()
            - recall_start
        )

        metric_time_totals[
            "recall"
        ] += recall_elapsed

        print(
            "  [TIME] "
            f"{'recall@1/3/5':22s}: "
            f"{recall_elapsed:8.4f}s"
        )

        # ====================================================
        # Recall 저장
        # ====================================================

        ws.cell(
            row=row,
            column=columns[
                "recall_at_1"
            ],
            value=r1,
        )

        ws.cell(
            row=row,
            column=columns[
                "recall_at_3"
            ],
            value=r3,
        )

        ws.cell(
            row=row,
            column=columns[
                "recall_at_5"
            ],
            value=r5,
        )

        # ====================================================
        # Unanswerable
        # ====================================================

        if r1 is None:

            ws.cell(
                row=row,
                column=recall_method_col,
                value=(
                    "N/A - "
                    "reference_text 없음"
                ),
            )

            ws.cell(
                row=row,
                column=recall_rank_col,
                value=None,
            )

            ws.cell(
                row=row,
                column=recall_score_col,
                value=None,
            )

            print(
                "  Recall@K       : N/A"
            )

        # ====================================================
        # Answerable
        # ====================================================

        else:

            answerable += 1

            recall_1_hits += (
                r1 or 0
            )

            recall_3_hits += (
                r3 or 0
            )

            recall_5_hits += (
                r5 or 0
            )

            # 가장 작은 K에서 성공한 결과 기록
            if r1 == 1:

                final_rank = rank1
                final_score = score1
                final_reason = reason1
                matched_scope = "Top-1"

            elif r3 == 1:

                final_rank = rank3
                final_score = score3
                final_reason = reason3
                matched_scope = "Top-3"

            elif r5 == 1:

                final_rank = rank5
                final_score = score5
                final_reason = reason5
                matched_scope = "Top-5"

            else:

                final_rank = None
                final_score = score5
                final_reason = reason5
                matched_scope = (
                    "Top-5 미탐지"
                )

            ws.cell(
                row=row,
                column=recall_method_col,
                value=(
                    "hybrid evidence match"
                    f" | scope={matched_scope}"
                    f" | reason={final_reason}"
                ),
            )

            ws.cell(
                row=row,
                column=recall_rank_col,
                value=final_rank,
            )

            ws.cell(
                row=row,
                column=recall_score_col,
                value=round(
                    final_score,
                    4,
                ),
            )

            # Console 출력
            print(
                "  Recall@1/3/5   : "
                f"{r1} / "
                f"{r3} / "
                f"{r5}"
            )

            print(
                "  Match Scope    : "
                f"{matched_scope}"
            )

            if final_rank is not None:

                print(
                    "  Match Rank     : "
                    f"{final_rank}"
                )

            elif (
                r1 == 1
                or r3 == 1
                or r5 == 1
            ):

                print(
                    "  Match Rank     : "
                    "복수 Context"
                )

            else:

                print(
                    "  Match Rank     : "
                    "없음"
                )

            print(
                "  Match Reason   : "
                f"{final_reason}"
            )

            print(
                "  Match Score    : "
                f"{final_score:.4f}"
            )

        # ====================================================
        # RAGAS
        # ====================================================

        if expected_behavior in {
            "refuse",
            "unanswerable",
        }:
            ws.cell(
                row=row,
                column=ragas_status_col,
                value="SKIPPED - UNANSWERABLE",
            )

        elif args.skip_ragas:
            current_status = ws.cell(
                row=row,
                column=ragas_status_col,
            ).value

            if current_status is None:
                ws.cell(
                    row=row,
                    column=ragas_status_col,
                    value="SKIPPED",
                )

        elif not response:
            ws.cell(
                row=row,
                column=ragas_status_col,
                value="SKIPPED - response 없음",
            )

        elif not contexts:
            ws.cell(
                row=row,
                column=ragas_status_col,
                value="SKIPPED - retrieved_contexts 없음",
            )

        else:
            assert scorers is not None

            run_faithfulness = (
                not args.factual_only
                and (
                    args.rerun_success
                    or not has_faithfulness
                )
            )

            run_factual = (
                args.rerun_success
                or not has_factual
            )

            if (
                not run_faithfulness
                and not run_factual
            ):
                print(
                    "  [RESUME] Faithfulness/Factual Correctness "
                    "이미 존재 → RAGAS SKIP"
                )

                ws.cell(
                    row=row,
                    column=ragas_status_col,
                    value="OK - RESUMED/SKIPPED",
                )

                if (
                    not args.factual_only
                    and has_faithfulness
                ):
                    try:
                        ragas_values[
                            "faithfulness"
                        ].append(
                            float(existing_faithfulness)
                        )
                    except (TypeError, ValueError):
                        pass

                if has_factual:
                    try:
                        ragas_values[
                            "factual_correctness"
                        ].append(
                            float(existing_factual)
                        )
                    except (TypeError, ValueError):
                        pass

            else:
                if (
                    not args.factual_only
                    and has_faithfulness
                    and not run_faithfulness
                ):
                    print(
                        "  [RESUME] faithfulness 기존 값 유지"
                    )

                if (
                    has_factual
                    and not run_factual
                ):
                    print(
                        "  [RESUME] factual_correctness 기존 값 유지"
                    )

                scores, metric_times = await score_one_with_ragas(
                    scorers=scorers,
                    user_input=user_input,
                    reference=reference,
                    response=response,
                    contexts=contexts,
                    factual_only=args.factual_only,
                    run_faithfulness=run_faithfulness,
                    run_factual_correctness=run_factual,
                )

                for metric_name, elapsed in metric_times.items():
                    metric_time_totals[
                        metric_name
                    ] += elapsed

                if run_faithfulness:
                    ws.cell(
                        row=row,
                        column=columns[
                            "faithfulness"
                        ],
                        value=scores[
                            "faithfulness"
                        ],
                    )

                if run_factual:
                    ws.cell(
                        row=row,
                        column=columns[
                            "factual_correctness"
                        ],
                        value=scores[
                            "factual_correctness"
                        ],
                    )

                final_faithfulness = (
                    scores["faithfulness"]
                    if run_faithfulness
                    else existing_faithfulness
                )

                final_factual = (
                    scores["factual_correctness"]
                    if run_factual
                    else existing_factual
                )

                if not args.factual_only:
                    try:
                        if (
                            final_faithfulness is not None
                            and str(final_faithfulness).strip() != ""
                        ):
                            ragas_values[
                                "faithfulness"
                            ].append(
                                float(final_faithfulness)
                            )
                    except (TypeError, ValueError):
                        pass

                try:
                    if (
                        final_factual is not None
                        and str(final_factual).strip() != ""
                    ):
                        ragas_values[
                            "factual_correctness"
                        ].append(
                            float(final_factual)
                        )
                except (TypeError, ValueError):
                    pass

                if args.factual_only:
                    status = (
                        "OK - FACTUAL_ONLY"
                        if final_factual is not None
                        and str(final_factual).strip() != ""
                        else "FAILED - FACTUAL_ONLY"
                    )
                elif (
                    final_faithfulness is not None
                    and str(final_faithfulness).strip() != ""
                    and final_factual is not None
                    and str(final_factual).strip() != ""
                ):
                    status = "OK"
                elif (
                    (
                        final_faithfulness is not None
                        and str(final_faithfulness).strip() != ""
                    )
                    or (
                        final_factual is not None
                        and str(final_factual).strip() != ""
                    )
                ):
                    status = "PARTIAL"
                else:
                    status = "FAILED"

                ws.cell(
                    row=row,
                    column=ragas_status_col,
                    value=status,
                )

        question_elapsed = (
            time.perf_counter()
            - question_start
        )

        print(
            "  [TIME] "
            f"{'QUESTION TOTAL':22s}: "
            f"{question_elapsed:8.2f}s"
        )

        # 중간 저장
        wb.save(
            output_path
        )

    # ========================================================
    # 선택 문항 확인
    # ========================================================

    if selected_question_ids is not None:

        missing_question_ids = (
            selected_question_ids
            - found_question_ids
        )

        if missing_question_ids:

            print(
                "\n[경고] Excel에서 "
                "찾지 못한 문항: "
                + ", ".join(
                    sorted(
                        missing_question_ids
                    )
                )
            )

    # ========================================================
    # 최종 Recall 통계
    # ========================================================

    print(
        "\n"
        + "=" * 78
    )

    print(
        "평가 완료"
    )

    print(
        f"전체 처리 질문 : "
        f"{processed}"
    )

    if answerable > 0:

        recall1 = (
            recall_1_hits
            / answerable
        )

        recall3 = (
            recall_3_hits
            / answerable
        )

        recall5 = (
            recall_5_hits
            / answerable
        )

        print(
            f"Answerable 질문 : "
            f"{answerable}"
        )

        print(
            "Recall@1        : "
            f"{recall_1_hits}/"
            f"{answerable} "
            f"= {recall1:.4f} "
            f"({recall1 * 100:.1f}%)"
        )

        print(
            "Recall@3        : "
            f"{recall_3_hits}/"
            f"{answerable} "
            f"= {recall3:.4f} "
            f"({recall3 * 100:.1f}%)"
        )

        print(
            "Recall@5        : "
            f"{recall_5_hits}/"
            f"{answerable} "
            f"= {recall5:.4f} "
            f"({recall5 * 100:.1f}%)"
        )

    # ========================================================
    # Correct Rejection Rate
    # ========================================================

    if unanswerable > 0:
        correct_rejection_rate = (
            correct_rejection_hits
            / unanswerable
        )

        print(
            "Correct Rejection : "
            f"{correct_rejection_hits}/"
            f"{unanswerable} "
            f"= {correct_rejection_rate:.4f} "
            f"({correct_rejection_rate * 100:.1f}%)"
        )

    # ========================================================
    # RAGAS 평균
    # ========================================================

    if not args.skip_ragas:

        print(
            "\n[RAGAS 평균]"
        )

        for (
            metric_name,
            values,
        ) in ragas_values.items():

            if (
                args.factual_only
                and metric_name
                != "factual_correctness"
            ):
                continue

            if values:

                average = (
                    sum(values)
                    / len(values)
                )

                print(
                    f"{metric_name:22s}: "
                    f"{average:.4f} "
                    f"({len(values)}개)"
                )

            else:

                print(
                    f"{metric_name:22s}: "
                    "N/A"
                )

    evaluation_elapsed = (
        time.perf_counter()
        - evaluation_start
    )

    print(
        "\n[실행시간 요약]"
    )

    print(
        f"{'Recall@1/3/5 합계':26s}: "
        f"{metric_time_totals['recall']:.2f}s"
    )

    if not args.skip_ragas:

        for metric_name in (
            "faithfulness",
            "factual_correctness",
        ):
            if (
                args.factual_only
                and metric_name
                != "factual_correctness"
            ):
                continue

            print(
                f"{metric_name:26s}: "
                f"{metric_time_totals[metric_name]:.2f}s"
            )

    print(
        f"{'전체 평가시간':26s}: "
        f"{evaluation_elapsed:.2f}s "
        f"({evaluation_elapsed / 60:.1f}분)"
    )

    if processed > 0:
        print(
            f"{'문항당 평균시간':26s}: "
            f"{evaluation_elapsed / processed:.2f}s"
        )

    print(
        f"\n결과 파일       : "
        f"{output_path}"
    )

    print(
        "=" * 78
    )


# ============================================================
# CLI
# ============================================================


def parse_args() -> argparse.Namespace:

    parser = argparse.ArgumentParser(
        description=(
            "evaluate_rag.py 결과를 읽어 "
            "Recall@K와 RAGAS 평가를 수행합니다."
        )
    )

    parser.add_argument(
        "--dataset",
        default="GC",
        help=(
            "평가셋 코드 "
            "(GC=고창율계, BD=서울번동3)"
        ),
    )

    parser.add_argument(
        "--xlsx",
        default=None,
        help=(
            "평가 입력 Excel 경로. "
            "생략하면 dataset 기본 경로 사용."
        ),
    )

    parser.add_argument(
        "--output",
        default=None,
        help=(
            "평가 결과 Excel 저장 경로. "
            "생략하면 dataset 기본 경로 사용."
        ),
    )

    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
    )

    parser.add_argument(
        "--skip-ragas",
        action="store_true",
        help=(
            "RAGAS 평가는 실행하지 않고 "
            "Recall@K만 다시 계산합니다."
        ),
    )

    parser.add_argument(
        "--ragas-base-url",
        default=DEFAULT_RAGAS_BASE_URL,
    )

    parser.add_argument(
        "--ragas-api-key",
        default=DEFAULT_RAGAS_API_KEY,
    )

    parser.add_argument(
        "--ragas-model",
        default=DEFAULT_RAGAS_MODEL,
    )


    parser.add_argument(
        "--factual-only",
        action="store_true",
        help=(
            "Faithfulness를 생략하고 "
            "Factual Correctness만 계산합니다. "
            "Judge 비교/디버깅용 빠른 평가 옵션입니다."
        ),
    )


    parser.add_argument(
        "--rerun-success",
        action="store_true",
        help=(
            "기존 출력 파일에 값이 있는 metric도 "
            "강제로 다시 계산합니다. "
            "기본값은 성공 metric 자동 SKIP입니다."
        ),
    )

    parser.add_argument(
        "--question-ids",
        default=None,
        help=(
            "평가할 question_id를 "
            "쉼표로 구분하여 지정합니다. "
            "예: Q002,Q003,Q007 "
            "생략하면 전체 문항을 평가합니다."
        ),
    )

    parser.add_argument(
        "--adapt-factual-korean",
        action="store_true",
        help=(
            "RAGAS FactualCorrectness의 "
            "claim 분해 prompt와 NLI prompt를 "
            "한국어로 adaptation합니다."
        ),
    )

    return parser.parse_args()


# ============================================================
# Main
# ============================================================


def main() -> None:

    args = parse_args()

    try:

        asyncio.run(
            evaluate_metrics(
                args
            )
        )

    except KeyboardInterrupt:

        print(
            "\n사용자가 평가를 "
            "중단했습니다."
        )

        sys.exit(130)

    except Exception as exc:

        print(
            "\n평가 중 오류가 "
            "발생했습니다:\n"
            f"{exc}"
        )

        sys.exit(1)


if __name__ == "__main__":
    main()