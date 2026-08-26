from __future__ import annotations

import argparse
import subprocess
import sys
import time
from pathlib import Path

from openpyxl import load_workbook

from evaluation.fixed_rag.service import answer_question


PROJECT_ROOT = Path(__file__).resolve().parents[1]

DATASET_DIR = (
    PROJECT_ROOT
    / "evaluation"
    / "datasets"
)

RESULT_DIR = (
    PROJECT_ROOT
    / "evaluation"
    / "results"
)


DEFAULT_SHEET_NAME = "평가셋"
DEFAULT_RUN_NUMBER = "001"


# ============================================================
# 고정 평가 Dataset ↔ 문서 ID
# ============================================================

FIXED_DOCUMENT_IDS = {
    "GC": "DOC_GC_001",
    "BD": "DOC_BD_001",
}


# ============================================================
# Dataset
# ============================================================

def normalize_dataset_name(
    dataset: str,
) -> str:

    normalized = (
        dataset
        .strip()
        .upper()
    )

    if not normalized:
        raise ValueError(
            "dataset이 비어 있습니다."
        )

    if normalized not in FIXED_DOCUMENT_IDS:
        raise ValueError(
            "지원하지 않는 고정 평가 dataset입니다: "
            f"{normalized}\n"
            "사용 가능: "
            + ", ".join(
                sorted(
                    FIXED_DOCUMENT_IDS
                )
            )
        )

    return normalized


def dataset_input_path(
    dataset: str,
) -> Path:

    path = (
        DATASET_DIR
        / f"{dataset}_FINAL_V1.xlsx"
    )

    if not path.is_file():
        raise FileNotFoundError(
            "평가셋 Excel을 찾을 수 없습니다: "
            f"{path}"
        )

    return path


def output_path(
    dataset: str,
    run_id: str,
) -> Path:

    return (
        RESULT_DIR
        / (
            f"{dataset}_FINAL_V1_"
            f"FIXED_RUN_{run_id}_result.xlsx"
        )
    )


# ============================================================
# Excel
# ============================================================

def find_columns(
    ws,
) -> dict[str, int]:

    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def require_columns(
    columns: dict[str, int],
    names: list[str],
) -> None:

    missing = [
        name
        for name in names
        if name not in columns
    ]

    if missing:
        raise ValueError(
            "필수 Excel 열이 없습니다: "
            + ", ".join(
                missing
            )
        )


def is_successful_response(
    value,
) -> bool:
    """
    이미 정상 답변이 저장된 행인지 확인한다.

    정상 답변:
        response가 비어 있지 않고
        [FIXED_RAG_ERROR]로 시작하지 않는 경우
    """

    text = str(
        value or ""
    ).strip()

    if not text:
        return False

    if text.startswith(
        "[FIXED_RAG_ERROR]"
    ):
        return False

    return True


# ============================================================
# Git
# ============================================================

def git_commit() -> str:

    try:
        return (
            subprocess.check_output(
                [
                    "git",
                    "rev-parse",
                    "--short",
                    "HEAD",
                ],
                cwd=PROJECT_ROOT,
                text=True,
            )
            .strip()
        )

    except Exception:
        return ""


# ============================================================
# Retrieval 결과 → Excel 문자열
# ============================================================

def format_evidence(
    evidence: list[dict],
) -> tuple[str, str]:

    chunk_ids: list[str] = []
    blocks: list[str] = []

    for rank, item in enumerate(
        evidence,
        start=1,
    ):
        chunk_id = str(
            item.get(
                "chunkId",
                "",
            )
        )

        section = str(
            item.get(
                "sectionTitle",
                "",
            )
            or ""
        )

        content = str(
            item.get(
                "content",
                "",
            )
            or ""
        )

        score = item.get(
            "score"
        )

        chunk_ids.append(
            chunk_id
        )

        try:
            score_text = (
                f"{float(score):.6f}"
            )

        except (
            TypeError,
            ValueError,
        ):
            score_text = ""

        blocks.append(
            "\n".join(
                [
                    (
                        f"[rank={rank} | "
                        f"chunkId={chunk_id} | "
                        f"section={section} | "
                        f"score={score_text}]"
                    ),
                    content,
                ]
            )
        )

    return (
        ", ".join(
            chunk_ids
        ),
        "\n\n---\n\n".join(
            blocks
        ),
    )


# ============================================================
# LLM 서버 장애 판별
# ============================================================

def is_fatal_llm_server_error(
    exc: Exception,
) -> bool:
    """
    이후 문항을 계속 실행해도 전부 실패할 가능성이 높은
    llama-server 장애인지 판별한다.

    예:
    - OOM으로 서버 프로세스 종료
    - Connection refused
    - Remote end closed connection
    - 모델 로딩 중 503
    """

    text = (
        f"{type(exc).__name__}: {exc}"
        .lower()
    )

    fatal_markers = [
        "connection refused",
        "remote end closed connection",
        "서버에 연결할 수 없습니다",
        "loading model",
        "status=503",
        "code=503",
        "connection reset",
    ]

    return any(
        marker in text
        for marker in fatal_markers
    )


# ============================================================
# CLI
# ============================================================

def parse_args() -> argparse.Namespace:

    parser = argparse.ArgumentParser(
        description=(
            "고정 평가 문서의 evaluation/outputs 산출물을 "
            "사용해 실제 RAG Generation 로직으로 "
            "답변을 수집합니다."
        )
    )

    parser.add_argument(
        "--dataset",
        required=True,
        help="GC 또는 BD",
    )

    parser.add_argument(
        "--xlsx",
        default=None,
        help=(
            "기본 평가셋이 아닌 다른 Excel을 "
            "사용할 경우 경로 지정"
        ),
    )

    parser.add_argument(
        "--output",
        default=None,
        help=(
            "결과 Excel 저장 경로를 "
            "직접 지정할 경우 사용"
        ),
    )

    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
    )

    parser.add_argument(
        "--run-id",
        default=DEFAULT_RUN_NUMBER,
    )

    parser.add_argument(
        "--top-k",
        type=int,
        default=None,
    )

    parser.add_argument(
        "--question-ids",
        default=None,
        help=(
            "일부 문항만 실행할 때 사용 "
            "예: Q001,Q002,Q010"
        ),
    )

    parser.add_argument(
        "--rerun-success",
        action="store_true",
        help=(
            "이미 정상 답변이 있는 문항도 "
            "다시 실행합니다."
        ),
    )

    return parser.parse_args()


# ============================================================
# Main
# ============================================================

def main() -> None:

    args = parse_args()

    dataset = normalize_dataset_name(
        args.dataset
    )

    document_id = (
        FIXED_DOCUMENT_IDS[
            dataset
        ]
    )

    # --------------------------------------------------------
    # 원본 평가셋
    # --------------------------------------------------------

    input_path = (
        Path(args.xlsx).resolve()
        if args.xlsx
        else dataset_input_path(
            dataset
        )
    )

    # --------------------------------------------------------
    # 결과 저장 경로
    # --------------------------------------------------------

    result_path = (
        Path(args.output).resolve()
        if args.output
        else output_path(
            dataset,
            args.run_id,
        )
    )

    if (
        args.top_k is not None
        and args.top_k <= 0
    ):
        raise ValueError(
            "--top-k는 1 이상이어야 합니다."
        )

    # --------------------------------------------------------
    # 특정 문항 선택
    # --------------------------------------------------------

    selected_question_ids = None

    if args.question_ids:
        selected_question_ids = {
            item.strip()
            for item
            in args.question_ids.split(",")
            if item.strip()
        }

    RESULT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # ========================================================
    # 중요:
    # 기존 결과 파일이 있으면 그 파일을 이어서 사용한다.
    # ========================================================

    if result_path.is_file():

        workbook_source = (
            result_path
        )

        print(
            "[RESUME] 기존 결과 파일을 이어서 사용합니다."
        )

    else:

        workbook_source = (
            input_path
        )

        print(
            "[NEW] 원본 평가셋에서 새 결과를 시작합니다."
        )

    wb = load_workbook(
        workbook_source
    )

    if args.sheet not in wb.sheetnames:
        raise ValueError(
            f"'{args.sheet}' 시트가 없습니다."
        )

    ws = wb[
        args.sheet
    ]

    columns = find_columns(
        ws
    )

    require_columns(
        columns,
        [
            "question_id",
            "user_input",
            "retrieved_chunk_ids",
            "retrieved_contexts",
            "response",
            "run_id",
            "git_commit",
        ],
    )

    current_commit = (
        git_commit()
    )

    processed = 0
    skipped = 0
    failed = 0

    server_stopped = False

    # --------------------------------------------------------
    # 실행 정보
    # --------------------------------------------------------

    print(
        "=" * 78
    )

    print(
        "고정 문서 RAG 결과 수집"
    )

    print(
        f"dataset     : {dataset}"
    )

    print(
        f"document_id : {document_id}"
    )

    print(
        f"input       : {input_path}"
    )

    print(
        f"workbook    : {workbook_source}"
    )

    print(
        f"output      : {result_path}"
    )

    print(
        f"run_id      : {args.run_id}"
    )

    print(
        "source      : "
        f"evaluation/source_documents/"
        f"{document_id}"
    )

    print(
        "=" * 78
    )

    # ========================================================
    # 질문 실행
    # ========================================================

    for row in range(
        2,
        ws.max_row + 1,
    ):

        question_id = str(
            ws.cell(
                row=row,
                column=columns[
                    "question_id"
                ],
            ).value
            or ""
        ).strip()

        if (
            selected_question_ids
            and question_id
            not in selected_question_ids
        ):
            continue

        question = str(
            ws.cell(
                row=row,
                column=columns[
                    "user_input"
                ],
            ).value
            or ""
        ).strip()

        if not question:
            continue

        # ====================================================
        # 이미 성공한 문항은 유지
        # ====================================================

        existing_response = (
            ws.cell(
                row=row,
                column=columns[
                    "response"
                ],
            ).value
        )

        if (
            not args.rerun_success
            and is_successful_response(
                existing_response
            )
        ):
            print(
                f"\n[{question_id}] "
                "이미 정상 답변 존재 → SKIP"
            )

            skipped += 1

            continue

        print(
            f"\n[{question_id}] "
            f"{question}"
        )

        started = (
            time.perf_counter()
        )

        # ====================================================
        # Fixed RAG 실행
        # ====================================================

        try:

            result = answer_question(
                dataset=dataset,
                question=question,
                top_k=args.top_k,
            )

            evidence = result.get(
                "evidence",
                [],
            )

            chunk_ids, contexts = (
                format_evidence(
                    evidence
                )
            )

            # 검색 Chunk
            ws.cell(
                row=row,
                column=columns[
                    "retrieved_chunk_ids"
                ],
                value=chunk_ids,
            )

            # 검색 Context
            ws.cell(
                row=row,
                column=columns[
                    "retrieved_contexts"
                ],
                value=contexts,
            )

            # 답변
            answer = str(
                result.get(
                    "answer",
                    "",
                )
            )

            ws.cell(
                row=row,
                column=columns[
                    "response"
                ],
                value=answer,
            )

            print(
                "  evidence : "
                f"{len(evidence)}개"
            )

            print(
                "  answer   : "
                f"{answer[:120]}"
            )

            processed += 1

        # ====================================================
        # 실패
        # ====================================================

        except Exception as exc:

            failed += 1

            error_text = (
                f"{type(exc).__name__}: "
                f"{exc}"
            )

            print(
                "  [실패] "
                f"{error_text}"
            )

            ws.cell(
                row=row,
                column=columns[
                    "response"
                ],
                value=(
                    "[FIXED_RAG_ERROR] "
                    f"{error_text}"
                ),
            )

            # -----------------------------------------------
            # 현재 실패 문항까지 즉시 저장
            # -----------------------------------------------

            wb.save(
                result_path
            )

            # -----------------------------------------------
            # llama-server 자체 장애면
            # 뒤 문항을 실행하지 않고 즉시 중단
            # -----------------------------------------------

            if is_fatal_llm_server_error(
                exc
            ):

                server_stopped = True

                print()
                print(
                    "=" * 78
                )
                print(
                    "[평가 중단]"
                )
                print(
                    "LLM 서버 연결 장애가 발생했습니다."
                )
                print(
                    "이후 문항은 실행하지 않습니다."
                )
                print(
                    "Gemma 서버를 재시작한 뒤 "
                    "같은 명령을 다시 실행하면 "
                    "성공한 문항은 자동으로 건너뜁니다."
                )
                print(
                    "=" * 78
                )

                break

        # ----------------------------------------------------
        # 수행 시간
        # ----------------------------------------------------

        elapsed_ms = int(
            (
                time.perf_counter()
                - started
            )
            * 1000
        )

        if "elapsed_ms" in columns:
            ws.cell(
                row=row,
                column=columns[
                    "elapsed_ms"
                ],
                value=elapsed_ms,
            )

        # Run ID
        ws.cell(
            row=row,
            column=columns[
                "run_id"
            ],
            value=(
                f"FIXED_RUN_"
                f"{args.run_id}"
            ),
        )

        # Git Commit
        ws.cell(
            row=row,
            column=columns[
                "git_commit"
            ],
            value=current_commit,
        )

        # 원본 문서
        if "source_file" in columns:
            ws.cell(
                row=row,
                column=columns[
                    "source_file"
                ],
                value=(
                    "evaluation/"
                    "source_documents/"
                    f"{document_id}"
                ),
            )

        # ====================================================
        # 질문 하나 끝날 때마다 즉시 저장
        # ====================================================

        wb.save(
            result_path
        )

    # ========================================================
    # 마지막 저장
    # ========================================================

    wb.save(
        result_path
    )

    print()

    print(
        "=" * 78
    )

    print(
        "고정 문서 RAG 결과 수집 종료"
    )

    print(
        f"신규 성공 : {processed}"
    )

    print(
        f"기존 성공 SKIP : {skipped}"
    )

    print(
        f"실패 : {failed}"
    )

    print(
        f"결과 파일 : {result_path}"
    )

    if server_stopped:
        print(
            "상태 : LLM 서버 장애로 중단"
        )
    else:
        print(
            "상태 : 정상 완료"
        )

    print(
        "=" * 78
    )


# ============================================================
# Entry Point
# ============================================================

if __name__ == "__main__":

    try:
        main()

    except KeyboardInterrupt:

        print(
            "\n사용자가 평가를 중단했습니다."
        )

        sys.exit(
            130
        )

    except Exception as exc:

        print(
            "\n고정 평가 실행 중 오류가 발생했습니다:\n"
            f"{exc}"
        )

        sys.exit(
            1
        )